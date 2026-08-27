#!/usr/bin/env python3
"""Continua a importação do Excel do Microsoft Forms para o GLPI sem duplicidade.

Também sincroniza o status de chamados antigos quando ele mudou na planilha.
"""

import argparse
import csv
import html
import json
import mimetypes
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import warnings
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", message="Cell .* is marked as a date.*", category=UserWarning)


STATUS_GLPI = {
    "novo": 1,                    # Novo
    "em andamento": 4,            # Pendente
    "aguardando validação": 4,    # Pendente
    "finalizado": 6,              # Fechado
    "suspenso": 6,                # Fechado
}

CATEGORIAS_CANONICAS = {
    "hardware (equipamento)": "Hardware (Equipamentos)",
    "hardware (equipamentos)": "Hardware (Equipamentos)",
    "solicitação de equipamento": "Solicitação de Equipamentos",
    "solicitação de equipamentos": "Solicitação de Equipamentos",
}

PRIORIDADE_GLPI = {
    "baixa": 2,
    "normal": 3,
    "urgente": 5,
}


# A coluna de data de abertura já apareceu com nomes diferentes em
# exportações diferentes do Microsoft Forms; qualquer um destes é aceito
# e tratado internamente como "Carimbo de data/hora".
ALIASES_COLUNA_ABERTURA = {"carimbo de data/hora", "data de abertura"}


def normalizar_colunas(df):
    renomear = {}
    for coluna in df.columns:
        if str(coluna).strip().casefold() in ALIASES_COLUNA_ABERTURA:
            renomear[coluna] = "Carimbo de data/hora"
    return df.rename(columns=renomear) if renomear else df


def texto(valor):
    if valor is None or pd.isna(valor):
        return ""
    return str(valor).strip()


def chave(valor):
    return texto(valor).casefold()


def linha_valida(linha):
    """Ignora linhas de formatação, espaços e controles sem um chamado real."""
    tem_data = bool(texto(linha.get("Carimbo de data/hora")))
    tem_solicitante = bool(texto(linha.get("Nome do Solicitante")))
    tem_descricao = bool(
        texto(linha.get("Descrição do Chamado"))
        or texto(linha.get("Motivo da Solicitação"))
    )
    return tem_data and tem_solicitante and tem_descricao


def valor_referencia(numero):
    achado = re.fullmatch(r"EXCEL-(\d+)", texto(numero), flags=re.IGNORECASE)
    if not achado:
        raise ValueError(f"Referência inválida: {numero}")
    return int(achado.group(1))


def data_glpi(valor):
    if valor is None or pd.isna(valor):
        return None
    try:
        dt = pd.to_datetime(valor, errors="raise", dayfirst=True)
    except Exception:
        return None
    if dt.year < 2023 or dt > pd.Timestamp.now() + pd.Timedelta(1, unit="D"):
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def carregar_config(caminho):
    with open(caminho, "r", encoding="utf-8") as arquivo:
        config = json.load(arquivo)
    obrigatorios = ["api_url", "app_token", "user_token"]
    ausentes = [campo for campo in obrigatorios if not texto(config.get(campo))]
    if ausentes:
        raise ValueError("Preencha no config.json: " + ", ".join(ausentes))
    config["api_url"] = config["api_url"].rstrip("/")
    return config


def montar_titulo(linha, numero):
    assunto = texto(linha.get("Assunto")) or "Chamado histórico"
    descricao = texto(linha.get("Descrição do Chamado")) or texto(linha.get("Motivo da Solicitação"))
    resumo = descricao.splitlines()[0][:90] if descricao else "Sem descrição"
    return f"[MIGRAÇÃO {numero}] {assunto} - {resumo}"[:250]


def montar_conteudo(linha, numero):
    nome = " ".join(filter(None, [texto(linha.get("Nome do Solicitante")), texto(linha.get("Sobrenome do Solicitante"))]))
    campos = [
        ("Referência da migração", numero),
        ("Solicitante original", nome),
        ("E-mail", texto(linha.get("Endereço de e-mail"))),
        ("Departamento", texto(linha.get("Departamento"))),
        ("Assunto/categoria original", texto(linha.get("Assunto"))),
        ("Tipo informado", texto(linha.get("Tipo"))),
        ("Motivo da solicitação", texto(linha.get("Motivo da Solicitação"))),
        ("Prioridade original", texto(linha.get("Prioridade"))),
        ("Quando o problema começou", texto(linha.get("Quando o problema iniciou?"))),
        ("AnyDesk informado", texto(linha.get("Insira seu acesso ao Anydesk"))),
        ("Responsável original", texto(linha.get("Responsável"))),
        ("Status original", texto(linha.get("Status"))),
        ("Anexo informado 1", texto(linha.get("Adicione uma foto do erro/problema"))),
        ("Anexo informado 2", texto(linha.get("Adicione uma foto do erro/problema.1"))),
    ]
    detalhes = "<br>".join(
        f"<strong>{html.escape(rotulo)}:</strong> {html.escape(str(valor))}"
        for rotulo, valor in campos if valor
    )
    descricao_original = (
        texto(linha.get("Descrição do Chamado"))
        or texto(linha.get("Motivo da Solicitação"))
        or "Sem descrição no arquivo original"
    )
    descricao = html.escape(descricao_original).replace("\n", "<br>")
    return f"{detalhes}<hr><strong>Descrição original:</strong><br>{descricao}"


class Glpi:
    def __init__(self, config):
        self.config = config
        self.headers = {
            "Content-Type": "application/json",
            "App-Token": config["app_token"],
            "Authorization": f"user_token {config['user_token']}",
        }
        self.session_token = None

    def requisicao(self, metodo, endpoint, dados=None, timeout=45):
        corpo = json.dumps(dados).encode("utf-8") if dados is not None else None
        requisicao = urllib.request.Request(
            f"{self.config['api_url']}/{endpoint.lstrip('/')}",
            data=corpo,
            headers=self.headers,
            method=metodo,
        )
        try:
            with urllib.request.urlopen(requisicao, timeout=timeout) as resposta:
                conteudo = resposta.read().decode("utf-8")
        except urllib.error.HTTPError as erro:
            resposta_erro = erro.read().decode("utf-8", errors="replace")[:1500]
            raise RuntimeError(
                f"Falha na API em {metodo} {endpoint}: HTTP {erro.code} - {resposta_erro}"
            ) from erro
        return json.loads(conteudo) if conteudo else None

    def iniciar(self):
        resposta = self.requisicao("GET", "initSession", timeout=30)
        self.session_token = resposta["session_token"]
        self.headers["Session-Token"] = self.session_token
        perfil_id = self.config.get("profile_id", 4)
        if perfil_id:
            self.requisicao("POST", "changeActiveProfile", {"profiles_id": int(perfil_id)})

    def finalizar(self):
        if self.session_token:
            try:
                self.requisicao("GET", "killSession", timeout=15)
            except Exception:
                pass

    def criar(self, tipo, dados):
        # Migração histórica: nunca deve disparar e-mail de notificação real
        # para requerente, técnico ou grupo. Os dois campos cobrem tanto os
        # itens "comuns" (Ticket, ITILSolution, Document_Item -> _disablenotifications)
        # quanto os itens de vínculo (Ticket_User, Group_Ticket -> _disablenotif).
        dados = {"_disablenotifications": 1, "_disablenotif": 1, **dados}
        retorno = self.requisicao("POST", tipo, {"input": dados})
        if isinstance(retorno, list):
            retorno = retorno[0]
        return int(retorno["id"])

    def listar(self, tipo, tamanho=9999):
        retorno = self.requisicao("GET", f"{tipo}?range=0-{tamanho}")
        return retorno if isinstance(retorno, list) else []

    def referencias_migradas(self):
        """Confere no próprio GLPI as referências já existentes."""
        referencias = {}
        limite = int(self.config.get("duplicate_scan_range", 10000))
        tamanho_pagina = int(self.config.get("duplicate_scan_page_size", 200))
        tamanho_pagina = max(20, min(tamanho_pagina, 500))
        for inicio in range(0, limite, tamanho_pagina):
            fim = min(inicio + tamanho_pagina - 1, limite - 1)
            pagina = self.requisicao("GET", f"Ticket?range={inicio}-{fim}")
            tickets = pagina if isinstance(pagina, list) else []
            for ticket in tickets:
                titulo = texto(ticket.get("name"))
                achado = re.search(r"\[MIGRAÇÃO\s+(EXCEL-\d+)\]", titulo, flags=re.IGNORECASE)
                if achado and ticket.get("id"):
                    referencias[achado.group(1).upper()] = int(ticket["id"])
            if len(tickets) < tamanho_pagina:
                break
        return referencias

    def enviar_documento(self, caminho):
        limite = "----GLPI" + uuid.uuid4().hex
        nome = caminho.name
        manifesto = json.dumps({"input": {"name": nome, "_filename": [nome]}}, ensure_ascii=False)
        partes = []

        def adicionar(texto_parte):
            partes.append(texto_parte.encode("utf-8"))

        adicionar(f"--{limite}\r\n")
        adicionar('Content-Disposition: form-data; name="uploadManifest"\r\n')
        adicionar("Content-Type: application/json\r\n\r\n")
        adicionar(manifesto + "\r\n")
        adicionar(f"--{limite}\r\n")
        adicionar(f'Content-Disposition: form-data; name="filename[0]"; filename="{nome}"\r\n')
        adicionar("Content-Type: application/octet-stream\r\n\r\n")
        partes.append(caminho.read_bytes())
        adicionar(f"\r\n--{limite}--\r\n")
        corpo = b"".join(partes)
        headers = {k: v for k, v in self.headers.items() if k.lower() != "content-type"}
        headers["Content-Type"] = f"multipart/form-data; boundary={limite}"
        requisicao = urllib.request.Request(
            f"{self.config['api_url']}/Document/",
            data=corpo,
            headers=headers,
            method="POST",
        )
        with urllib.request.urlopen(requisicao, timeout=120) as resposta:
            retorno = json.loads(resposta.read().decode("utf-8"))
        return int(retorno["id"])

    def atualizar(self, tipo, item_id, dados):
        # Mesmo motivo do criar(): atualização de chamado antigo (status,
        # solução, categoria etc.) não deve notificar ninguém de verdade.
        self.requisicao(
            "PUT",
            f"{tipo}/{item_id}",
            {"input": {"id": item_id, "_disablenotifications": 1, "_disablenotif": 1, **dados}},
        )

    def obter_ou_none(self, tipo, item_id):
        try:
            retorno = self.requisicao("GET", f"{tipo}/{int(item_id)}")
        except RuntimeError as erro:
            mensagem = str(erro)
            if "HTTP 404" in mensagem or "ITEM_NOT_FOUND" in mensagem:
                return None
            raise
        if not isinstance(retorno, dict) or int(retorno.get("id") or 0) != int(item_id):
            return None
        return retorno


def chave_duplicidade(linha):
    """Impressão digital de um chamado: mesmo solicitante + mesma descrição."""
    nome = " ".join(filter(None, [
        texto(linha.get("Nome do Solicitante")),
        texto(linha.get("Sobrenome do Solicitante")),
    ]))
    descricao = texto(linha.get("Descrição do Chamado")) or texto(linha.get("Motivo da Solicitação"))
    descricao_normalizada = re.sub(r"\s+", " ", descricao).strip().casefold()
    return (chave(nome), descricao_normalizada)


def identificar_duplicados(novos):
    """Agrupa os chamados novos com a mesma impressão digital (mesmo
    solicitante e mesma descrição). Devolve {numero: numero_original} só
    para as repetições; a primeira ocorrência de cada grupo (a de menor
    EXCEL-N) continua sendo criada normalmente.
    """
    grupos = defaultdict(list)
    for _, numero, linha, _ in novos:
        chave_grupo = chave_duplicidade(linha)
        if not chave_grupo[1]:
            continue
        grupos[chave_grupo].append(numero)

    duplicados = {}
    for numeros in grupos.values():
        if len(numeros) < 2:
            continue
        numeros_ordenados = sorted(numeros, key=valor_referencia)
        original = numeros_ordenados[0]
        for numero in numeros_ordenados[1:]:
            duplicados[numero] = original
    return duplicados


def carregar_estado(caminho):
    if not caminho.exists():
        return {"importados": {}}
    with open(caminho, "r", encoding="utf-8") as arquivo:
        return json.load(arquivo)


def salvar_estado(caminho, estado):
    temporario = caminho.with_suffix(".tmp")
    with open(temporario, "w", encoding="utf-8") as arquivo:
        json.dump(estado, arquivo, ensure_ascii=False, indent=2)
    temporario.replace(caminho)


def categoria_canonica(valor):
    original = texto(valor) or "Outros"
    return CATEGORIAS_CANONICAS.get(chave(original), original)


def nomes_do_usuario(usuario):
    login = texto(usuario.get("name"))
    primeiro = texto(usuario.get("firstname"))
    sobrenome = texto(usuario.get("realname"))
    completos = {
        chave(login),
        chave(primeiro),
        chave(sobrenome),
        chave(" ".join(filter(None, [primeiro, sobrenome]))),
        chave(" ".join(filter(None, [sobrenome, primeiro]))),
    }
    return {nome for nome in completos if nome}


def preparar_categorias(glpi, config, df):
    categorias = glpi.listar("ITILCategory")
    categorias_por_nome = {
        chave(item.get("completename") or item.get("name")): int(item["id"])
        for item in categorias if item.get("id")
    }
    mapa_categorias = {chave(k): int(v) for k, v in config.get("category_map", {}).items()}

    for assunto in sorted({categoria_canonica(v) for v in df["Assunto"].tolist()}):
        normalizado = chave(assunto)
        if normalizado in mapa_categorias:
            continue
        existente = categorias_por_nome.get(normalizado)
        if existente:
            mapa_categorias[normalizado] = existente
            continue
        if config.get("auto_create_categories", False):
            novo_id = glpi.criar("ITILCategory", {"name": assunto, "is_helpdeskvisible": 1})
            mapa_categorias[normalizado] = novo_id
            print(f"Categoria criada: {assunto} (ID {novo_id})")
        else:
            print(f"AVISO: categoria não encontrada e não criada automaticamente: {assunto}")

    return mapa_categorias


def unico(conjunto):
    return next(iter(conjunto)) if conjunto and len(conjunto) == 1 else None


def registrar_pendencia(caminho, referencia, tipo, valor):
    novo = not caminho.exists()
    with open(caminho, "a", encoding="utf-8-sig", newline="") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        if novo:
            escritor.writerow(["Referência", "Tipo", "Valor"])
        escritor.writerow([referencia, tipo, valor])


def carregar_links_anexos(caminho_excel):
    planilha = load_workbook(caminho_excel, read_only=False, data_only=False).active
    anexos = {}
    for linha in range(2, planilha.max_row + 1):
        itens = []
        for coluna in (12, 14):
            celula = planilha.cell(linha, coluna)
            if celula.hyperlink and celula.value:
                itens.append((texto(celula.value), celula.hyperlink.target))
        if itens:
            anexos[linha] = itens
    return anexos


def baixar_google_drive(nome, url, pasta):
    consulta = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    arquivo_id = (consulta.get("id") or [None])[0]
    if not arquivo_id:
        partes = urllib.parse.urlparse(url).path.split("/")
        arquivo_id = partes[partes.index("d") + 1] if "d" in partes else None
    if not arquivo_id:
        raise ValueError("Link do Google Drive sem ID")
    download = "https://drive.usercontent.google.com/download?" + urllib.parse.urlencode({
        "id": arquivo_id, "export": "download", "confirm": "t"
    })
    requisicao = urllib.request.Request(download, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(requisicao, timeout=120) as resposta:
        tipo = resposta.headers.get("Content-Type", "")
        dados = resposta.read()
    if "text/html" in tipo.lower() or dados[:50].lower().find(b"<html") >= 0:
        raise PermissionError("Google Drive solicitou autenticação")
    nome_original = Path(nome).name if not urllib.parse.urlparse(nome).scheme else ""
    disposicao = resposta.headers.get("Content-Disposition", "")
    achado = re.search(r"filename\*?=(?:UTF-8''|\")?([^\";]+)", disposicao, flags=re.IGNORECASE)
    if achado:
        nome_original = urllib.parse.unquote(achado.group(1).strip())
    if not nome_original:
        extensao = mimetypes.guess_extension(tipo.split(";", 1)[0].strip()) or ""
        nome_original = f"anexo_{arquivo_id}{extensao}"
    nome_seguro = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", nome_original).strip(" .")
    destino = pasta / (nome_seguro or f"anexo_{arquivo_id}")
    destino.write_bytes(dados)
    return destino


def localizar_anexo_local(nome, pasta_raiz):
    if not pasta_raiz.exists():
        return None
    procurado = Path(nome).name.casefold()
    encontrados = [p for p in pasta_raiz.rglob("*") if p.is_file() and p.name.casefold() == procurado]
    return encontrados[0] if encontrados else None


def registrar_fechamento_sql(caminho, ticket_id, data_final):
    with open(caminho, "a", encoding="utf-8") as arquivo:
        arquivo.write(
            f"UPDATE glpi_tickets SET solvedate='{data_final}', closedate='{data_final}' WHERE id={int(ticket_id)};\n"
        )


def detectar_reparos(df_atual, df_anterior, estado):
    """Encontra referências antes vazias que foram preenchidas depois da 1ª importação."""
    reparos = []
    if df_anterior is None:
        return reparos
    for indice, linha in df_atual.iterrows():
        numero = f"EXCEL-{indice + 2}"
        controle = estado.get("importados", {}).get(numero)
        if not controle or controle.get("reparado_em") or not linha_valida(linha):
            continue
        linha_antiga = df_anterior.loc[indice] if indice in df_anterior.index else None
        if linha_antiga is None or not linha_valida(linha_antiga):
            reparos.append((indice, numero, linha, int(controle["ticket_id"])))
    return reparos


def detectar_status_alterados(df_atual, df_anterior, estado, numero_maximo):
    """Encontra chamados cujo status atual da planilha diverge do último
    status sincronizado com o GLPI (registrado no controle).

    Compara direto contra o controle em vez de comparar apenas a planilha
    atual com a anterior: uma comparação só entre as duas últimas planilhas
    perde mudanças de status que aconteceram e já não aparecem mais como
    diferença entre exatamente essas duas versões. Também não filtra por
    número máximo de referência: um chamado antigo pode ter o status da
    planilha alterado a qualquer momento, não só os mais recentes.
    """
    alterados = []
    for indice, linha in df_atual.iterrows():
        numero = f"EXCEL-{indice + 2}"
        if not linha_valida(linha):
            continue
        controle = estado.get("importados", {}).get(numero)
        if not controle or not controle.get("ticket_id"):
            continue
        status_atual = chave(linha.get("Status"))
        if not status_atual:
            continue
        status_final_atual = STATUS_GLPI.get(status_atual, 1)
        if controle.get("status_sincronizado_excel") == status_final_atual:
            continue
        alterados.append((indice, numero, linha, int(controle["ticket_id"])))
    return alterados


def main():
    parser = argparse.ArgumentParser(description="Importador incremental Excel Forms para GLPI")
    parser.add_argument("excel", help="Caminho do arquivo .xlsx")
    parser.add_argument("--config", default="config.json", help="Arquivo de configuração")
    parser.add_argument(
        "--excel-anterior",
        default="planilha_anterior.xlsx",
        help="Planilha usada na importação anterior; permite reparar linhas que antes estavam vazias",
    )
    parser.add_argument(
        "--depois-de",
        type=int,
        default=0,
        help="Só cria referências maiores que este número (padrão: 0, ou seja, todas)",
    )
    parser.add_argument("--limit", type=int, default=0, help="Máximo de chamados; 0 significa sem limite (padrão: 0)")
    parser.add_argument("--executar", action="store_true", help="Cria os chamados; sem isso apenas simula")
    parser.add_argument(
        "--reparar-preenchidos",
        action="store_true",
        help="Atualiza no GLPI os chamados já importados cuja linha estava vazia e foi preenchida depois",
    )
    parser.add_argument(
        "--sincronizar-status-antigos",
        action="store_true",
        help="Atualiza no GLPI os chamados antigos cujo status mudou na planilha atual",
    )
    parser.add_argument("--somente-anexo-ticket", type=int, help="Anexa arquivos a um chamado existente")
    parser.add_argument("--linha-excel", type=int, help="Número da linha do Excel usado com --somente-anexo-ticket")
    parser.add_argument("--reprocessar-anexos-pendentes", action="store_true", help="Tenta novamente os anexos listados como pendentes")
    parser.add_argument(
        "--permitir-duplicados",
        action="store_true",
        help="Não agrupa chamados repetidos (mesmo solicitante e mesma descrição); cria um chamado por linha",
    )
    args = parser.parse_args()

    excel = Path(args.excel).resolve()
    config_path = Path(args.config).resolve()
    excel_anterior = Path(args.excel_anterior).resolve() if args.excel_anterior else None
    if not excel.exists():
        raise FileNotFoundError(f"Excel não encontrado: {excel}")

    df = pd.read_excel(excel)
    df = normalizar_colunas(df)
    df_anterior = pd.read_excel(excel_anterior) if excel_anterior and excel_anterior.exists() else None
    if df_anterior is not None:
        df_anterior = normalizar_colunas(df_anterior)
    links_anexos = carregar_links_anexos(excel)
    if "Nome do Solicitante" not in df.columns:
        raise ValueError("A planilha não possui as colunas esperadas.")
    df = df[df["Nome do Solicitante"].fillna("").astype(str).str.strip().str.casefold() != "ignorar"]

    config = carregar_config(config_path) if args.executar else {}
    estado_path = Path("importacao_glpi_estado.json").resolve()
    estado = carregar_estado(estado_path)
    reparos = detectar_reparos(df, df_anterior, estado)
    status_alterados = detectar_status_alterados(
        df,
        df_anterior,
        estado,
        args.depois_de,
    )
    novos = []

    for indice, linha in df.iterrows():
        numero = f"EXCEL-{indice + 2}"
        if not linha_valida(linha):
            continue
        if numero in estado["importados"]:
            continue
        if valor_referencia(numero) <= args.depois_de:
            continue
        novos.append((indice, numero, linha, None))

    duplicados_map = {} if args.permitir_duplicados else identificar_duplicados(novos)

    selecionados = []
    if args.sincronizar_status_antigos:
        selecionados.extend(("SINCRONIZAR_STATUS", *item) for item in status_alterados)
    if args.reparar_preenchidos:
        selecionados.extend(("ATUALIZAR", *item) for item in reparos)
    if not args.sincronizar_status_antigos:
        for indice, numero, linha, ticket_existente in novos:
            if numero in duplicados_map:
                # ticket_existente aqui carrega a referência EXCEL-N original,
                # não um ticket_id: é resolvido no momento da criação.
                selecionados.append(("DUPLICADO", indice, numero, linha, duplicados_map[numero]))
            else:
                selecionados.append(("CRIAR", indice, numero, linha, ticket_existente))
    if args.limit > 0:
        selecionados = selecionados[:args.limit]

    if reparos and not args.reparar_preenchidos and not args.sincronizar_status_antigos:
        print(
            f"ATENÇÃO: {len(reparos)} referências importadas quando estavam vazias precisam de reparo: "
            + ", ".join(numero for _, numero, _, _ in reparos)
        )
        if args.executar:
            raise ValueError(
                "Execução bloqueada. Revise a simulação e acrescente --reparar-preenchidos "
                "para atualizar os chamados vazios sem criar duplicatas."
            )

    if duplicados_map:
        exemplos = ", ".join(
            f"{numero}=repetição de {original}"
            for numero, original in list(duplicados_map.items())[:10]
        )
        resto = len(duplicados_map) - 10
        if resto > 0:
            exemplos += f" (+{resto} outros)"
        print(
            f"REPETIDOS: {len(duplicados_map)} chamado(s) novo(s) têm o mesmo solicitante e a mesma "
            f"descrição de outra linha já selecionada; não vão criar chamado novo no GLPI, só serão "
            f"vinculados ao chamado original. Use --permitir-duplicados para forçar a criação de todos: {exemplos}"
        )

    if not args.somente_anexo_ticket and not args.reprocessar_anexos_pendentes:
        substantivo = "ação selecionada" if len(selecionados) == 1 else "ações selecionadas"
        print(
            f"PLANO: {len(status_alterados)} status antigos alterados; "
            f"{len(reparos)} reparos identificados; {len(novos)} chamados novos "
            f"({len(duplicados_map)} repetidos, não criarão chamado novo); "
            f"{len(selecionados)} {substantivo} neste comando."
        )

    if not selecionados and not args.somente_anexo_ticket and not args.reprocessar_anexos_pendentes:
        print("Nenhum chamado novo ou status antigo alterado encontrado.")
        return

    glpi = Glpi(config) if args.executar else None
    if glpi:
        glpi.iniciar()
        consultar_servidor = bool(config.get("scan_existing_tickets", False))
        if consultar_servidor:
            referencias_glpi = glpi.referencias_migradas()
        else:
            referencias_glpi = {}
            print(
                "DUPLICIDADE: usando importacao_glpi_estado.json e corte após "
                f"EXCEL-{args.depois_de}; consulta completa ao servidor desativada."
            )
        filtrados = []
        for acao, indice, numero, linha, ticket_existente in selecionados:
            encontrado = referencias_glpi.get(numero)
            if acao == "SINCRONIZAR_STATUS":
                ticket_atual = glpi.obter_ou_none("Ticket", ticket_existente)
                if ticket_atual is None:
                    print(
                        f"AUSENTE: {numero} aponta para GLPI #{ticket_existente}; "
                        "a sincronização de status foi ignorada e nenhum chamado foi criado."
                    )
                    continue
                status_desejado = STATUS_GLPI.get(chave(linha.get("Status")), 1)
                status_no_glpi = int(ticket_atual.get("status") or 0)
                if status_no_glpi == status_desejado:
                    estado["importados"][numero]["status_sincronizado_excel"] = status_desejado
                    estado["importados"][numero]["status_sincronizado_em"] = datetime.now().isoformat(timespec="seconds")
                    salvar_estado(estado_path, estado)
                    print(
                        f"JÁ OK: {numero} / GLPI #{ticket_existente} já está com "
                        f"status {status_desejado}."
                    )
                    continue
                filtrados.append((acao, indice, numero, linha, ticket_existente))
                continue
            if acao == "ATUALIZAR":
                ticket_atual = glpi.obter_ou_none("Ticket", ticket_existente)
                if ticket_atual is None:
                    print(
                        f"AUSENTE: {numero} apontava para GLPI #{ticket_existente}, "
                        "mas o item não existe; o chamado real será criado."
                    )
                    filtrados.append(("RECRIAR", indice, numero, linha, ticket_existente))
                    continue
                if encontrado and encontrado != ticket_existente:
                    raise RuntimeError(
                        f"{numero} aponta para GLPI #{ticket_existente} no controle, "
                        f"mas foi encontrado como GLPI #{encontrado} no servidor."
                    )
                filtrados.append((acao, indice, numero, linha, ticket_existente))
                continue
            if encontrado:
                estado["importados"][numero] = {
                    "ticket_id": encontrado,
                    "reconciliado_em": datetime.now().isoformat(timespec="seconds"),
                }
                salvar_estado(estado_path, estado)
                print(f"PULADO: {numero} já existe no GLPI #{encontrado}.")
                continue
            filtrados.append((acao, indice, numero, linha, ticket_existente))
        selecionados = filtrados

    pasta_temporaria = Path("anexos_temporarios").resolve()
    pasta_temporaria.mkdir(exist_ok=True)
    pasta_anexos_locais = Path(config.get("local_attachment_dir", "anexos_forms")).resolve() if args.executar else Path("anexos_forms")

    if args.reprocessar_anexos_pendentes:
        if not glpi:
            raise ValueError("Use --executar junto com --reprocessar-anexos-pendentes")
        pendencias_origem = Path("importacao_glpi_pendencias.csv").resolve()
        pendencias_novas = Path("importacao_glpi_pendencias_reprocessamento.csv").resolve()
        if pendencias_novas.exists():
            pendencias_novas.unlink()
        if not pendencias_origem.exists():
            print("Nenhuma pendência de anexo encontrada.")
            glpi.finalizar()
            return
        with open(pendencias_origem, "r", encoding="utf-8-sig", newline="") as arquivo:
            registros = list(csv.DictReader(arquivo, delimiter=";"))
        sucessos = 0
        falhas = 0
        try:
            for registro in registros:
                referencia = texto(registro.get("Referência"))
                if registro.get("Tipo") != "Anexo" or not referencia.startswith("EXCEL-"):
                    continue
                linha_excel = int(referencia.split("-", 1)[1])
                controle = estado.get("importados", {}).get(referencia)
                if not controle:
                    registrar_pendencia(pendencias_novas, referencia, "Anexo", "Chamado não encontrado no controle")
                    falhas += 1
                    continue
                ticket_id = int(controle["ticket_id"])
                linha_df = df.loc[df.index == linha_excel - 2]
                if linha_df.empty:
                    registrar_pendencia(pendencias_novas, referencia, "Anexo", "Linha não encontrada no Excel")
                    falhas += 1
                    continue
                status_original = chave(linha_df.iloc[0].get("Status"))
                status_final = STATUS_GLPI.get(status_original, 1)
                valor_pendente = texto(registro.get("Valor"))
                candidatos = [
                    (nome, url) for nome, url in links_anexos.get(linha_excel, [])
                    if valor_pendente.startswith(f"{nome} -")
                ]
                if not candidatos:
                    registrar_pendencia(pendencias_novas, referencia, "Anexo", "Link pendente não identificado")
                    falhas += 1
                    continue
                reaberto = status_final in {5, 6}
                try:
                    if reaberto:
                        glpi.atualizar("Ticket", ticket_id, {"status": 1})
                    for nome_anexo, url_anexo in candidatos:
                        caminho_local = localizar_anexo_local(nome_anexo, pasta_anexos_locais)
                        caminho_anexo = caminho_local or baixar_google_drive(nome_anexo, url_anexo, pasta_temporaria)
                        documento_id = glpi.enviar_documento(caminho_anexo)
                        glpi.criar("Document_Item", {
                            "documents_id": documento_id,
                            "items_id": ticket_id,
                            "itemtype": "Ticket",
                        })
                        if caminho_anexo.parent == pasta_temporaria:
                            caminho_anexo.unlink(missing_ok=True)
                    sucessos += 1
                    print(f"ANEXO OK: {referencia} no GLPI #{ticket_id}")
                except Exception as erro:
                    registrar_pendencia(pendencias_novas, referencia, "Anexo", str(erro))
                    falhas += 1
                finally:
                    if reaberto:
                        try:
                            glpi.atualizar("Ticket", ticket_id, {"status": status_final})
                        except Exception as erro_status:
                            registrar_pendencia(pendencias_novas, referencia, "Status", str(erro_status))
                            falhas += 1
                time.sleep(float(config.get("interval_seconds", 0.3)))
        finally:
            glpi.finalizar()
        print(f"REPROCESSAMENTO CONCLUÍDO: {sucessos} anexos OK; {falhas} falhas")
        return

    if args.somente_anexo_ticket:
        if not glpi or not args.linha_excel:
            raise ValueError("Use --executar e --linha-excel junto com --somente-anexo-ticket")
        try:
            for nome_anexo, url_anexo in links_anexos.get(args.linha_excel, []):
                caminho_local = localizar_anexo_local(nome_anexo, pasta_anexos_locais)
                caminho_anexo = caminho_local or baixar_google_drive(nome_anexo, url_anexo, pasta_temporaria)
                try:
                    documento_id = glpi.enviar_documento(caminho_anexo)
                except urllib.error.HTTPError as erro:
                    corpo = erro.read().decode("utf-8", errors="replace")[:1500]
                    raise RuntimeError(f"Falha no upload do documento: HTTP {erro.code} - {corpo}") from erro
                try:
                    glpi.criar("Document_Item", {
                        "documents_id": documento_id,
                        "items_id": args.somente_anexo_ticket,
                        "itemtype": "Ticket",
                    })
                except urllib.error.HTTPError as erro:
                    corpo = erro.read().decode("utf-8", errors="replace")[:1500]
                    raise RuntimeError(f"Documento {documento_id} enviado, mas falhou o vínculo: HTTP {erro.code} - {corpo}") from erro
                print(f"ANEXO OK: {nome_anexo} no chamado #{args.somente_anexo_ticket}")
        finally:
            glpi.finalizar()
        return

    mapa_categorias = {}
    if glpi:
        df_selecionado = pd.DataFrame([linha for _, _, _, linha, _ in selecionados])
        mapa_categorias = preparar_categorias(glpi, config, df_selecionado)
    pendencias_path = Path("importacao_glpi_pendencias.csv").resolve()
    sql_fechamentos_path = Path("corrigir_datas_fechamento.sql").resolve()

    try:
        for acao, _, numero, linha, ticket_existente in selecionados:
            status_original = chave(linha.get("Status"))
            prioridade = PRIORIDADE_GLPI.get(chave(linha.get("Prioridade")), 3)
            data_abertura = data_glpi(linha.get("Carimbo de data/hora"))
            data_final = data_glpi(linha.get("Data da Finalização"))

            if acao == "SINCRONIZAR_STATUS":
                status_final = STATUS_GLPI.get(status_original, 1)
                if not args.executar:
                    print(json.dumps({
                        "acao": acao,
                        "referencia": numero,
                        "ticket_id_existente": ticket_existente,
                        "status_excel": texto(linha.get("Status")),
                        "status_glpi_desejado": status_final,
                        "data_finalizacao": data_final,
                    }, ensure_ascii=False, indent=2))
                    continue

                ticket_id = int(ticket_existente)
                try:
                    ticket_atual = glpi.requisicao("GET", f"Ticket/{ticket_id}")
                    status_glpi_atual = int(ticket_atual.get("status") or 0)
                    if status_glpi_atual == 6 and status_final != 6:
                        # Chamado fechado no GLPI: reabre antes de editar
                        # (mesmo motivo do reparo de referências preenchidas).
                        glpi.atualizar("Ticket", ticket_id, {"status": 1})

                    solucao = texto(linha.get("Desdobramento"))
                    if status_final == 6 and not solucao:
                        solucao = "Chamado concluído no controle antigo; a solução não foi registrada no Excel."
                    if status_final == 6 and solucao:
                        solucoes = []
                        try:
                            retorno_solucoes = glpi.requisicao(
                                "GET",
                                f"Ticket/{ticket_id}/ITILSolution?range=0-99",
                            )
                            solucoes = retorno_solucoes if isinstance(retorno_solucoes, list) else []
                        except Exception:
                            # A consulta de subitens pode não estar disponível em todos
                            # os perfis. Nesse caso, registra a solução necessária.
                            solucoes = []
                        if not solucoes:
                            glpi.criar("ITILSolution", {
                                "itemtype": "Ticket",
                                "items_id": ticket_id,
                                "content": html.escape(solucao).replace("\n", "<br>"),
                            })
                    if data_final and status_final == 6:
                        glpi.atualizar("Ticket", ticket_id, {"solvedate": data_final})
                    glpi.atualizar("Ticket", ticket_id, {"status": status_final})
                    if data_final and status_original == "finalizado":
                        registrar_fechamento_sql(sql_fechamentos_path, ticket_id, data_final)

                    estado["importados"][numero]["status_sincronizado_excel"] = status_final
                    estado["importados"][numero]["status_sincronizado_em"] = datetime.now().isoformat(timespec="seconds")
                    salvar_estado(estado_path, estado)
                    print(
                        f"OK STATUS: {numero} atualizado no GLPI #{ticket_id} "
                        f"de {status_glpi_atual} para {status_final} "
                        f"({texto(linha.get('Status'))})."
                    )
                except Exception as erro_status:
                    registrar_pendencia(pendencias_path, numero, "Status", str(erro_status))
                    print(f"ERRO STATUS: {numero} / GLPI #{ticket_id} | {erro_status}")
                time.sleep(float(config.get("interval_seconds", 0.3)))
                continue

            if acao == "DUPLICADO":
                referencia_original = ticket_existente
                controle_original = estado["importados"].get(referencia_original)
                ticket_id_original = (
                    int(controle_original["ticket_id"])
                    if controle_original and controle_original.get("ticket_id")
                    else None
                )
                if not args.executar:
                    print(json.dumps({
                        "acao": acao,
                        "referencia": numero,
                        "duplicado_de": referencia_original,
                        "ticket_id_original": ticket_id_original,
                    }, ensure_ascii=False, indent=2))
                    continue
                if not ticket_id_original:
                    print(
                        f"AGUARDANDO: {numero} é repetição de {referencia_original}, que ainda não foi "
                        "criado nesta execução (provavelmente ficou fora do --limit); será resolvido "
                        "em uma próxima rodada, sem criar chamado duplicado."
                    )
                    continue
                # Os anexos da linha repetida podem ser fotos diferentes do
                # mesmo problema; são vinculados ao chamado original.
                for nome_anexo, url_anexo in links_anexos.get(int(numero.split("-")[1]), []):
                    try:
                        caminho_local = localizar_anexo_local(nome_anexo, pasta_anexos_locais)
                        caminho_anexo = caminho_local or baixar_google_drive(nome_anexo, url_anexo, pasta_temporaria)
                        documento_id = glpi.enviar_documento(caminho_anexo)
                        glpi.criar("Document_Item", {
                            "documents_id": documento_id,
                            "items_id": ticket_id_original,
                            "itemtype": "Ticket",
                        })
                        if caminho_anexo.parent == pasta_temporaria:
                            caminho_anexo.unlink(missing_ok=True)
                    except Exception as erro_anexo:
                        registrar_pendencia(pendencias_path, numero, "Anexo", f"{nome_anexo} - {erro_anexo}")

                estado["importados"][numero] = {
                    "ticket_id": ticket_id_original,
                    "duplicado_de": referencia_original,
                    "vinculado_em": datetime.now().isoformat(timespec="seconds"),
                }
                salvar_estado(estado_path, estado)
                print(
                    f"DUPLICADO: {numero} tem o mesmo solicitante e a mesma descrição de "
                    f"{referencia_original}; não criou chamado novo, foi vinculado ao GLPI #{ticket_id_original}."
                )
                time.sleep(float(config.get("interval_seconds", 0.3)))
                continue

            ticket = {
                "name": montar_titulo(linha, numero),
                "content": montar_conteudo(linha, numero),
                "status": 1,
                "priority": prioridade,
                "urgency": prioridade,
                "type": 2,
            }
            if data_abertura:
                ticket["date"] = data_abertura

            categoria_id = mapa_categorias.get(chave(categoria_canonica(linha.get("Assunto")))) if args.executar else None
            if categoria_id:
                ticket["itilcategories_id"] = int(categoria_id)

            responsavel = texto(linha.get("Responsável"))
            tecnico_id = config.get("technician_map", {}).get(responsavel) if args.executar else None
            tecnico_id = int(tecnico_id) if tecnico_id else None
            if tecnico_id:
                ticket["_users_id_assign"] = int(tecnico_id)

            if not args.executar:
                print(json.dumps({
                    "acao": acao,
                    "referencia": numero,
                    "ticket_id_existente": ticket_existente,
                    "ticket": ticket,
                    "status_final": STATUS_GLPI.get(status_original, 1),
                }, ensure_ascii=False, indent=2))
                continue

            if acao == "ATUALIZAR":
                ticket_id = int(ticket_existente)
                # O GLPI rejeita edição de campos em chamados já fechados/solucionados;
                # reabre antes de editar. O status correto é restaurado no final do laço.
                glpi.atualizar("Ticket", ticket_id, {"status": 1})
                # Atualizações de chamados existentes são divididas porque o
                # GLPI rejeita campos auxiliares de criação em requisições PUT.
                glpi.atualizar("Ticket", ticket_id, {
                    "name": ticket["name"],
                    "content": ticket["content"],
                })
                campos_operacionais = {
                    campo: ticket[campo]
                    for campo in ("date", "priority", "urgency", "type")
                    if campo in ticket
                }
                if campos_operacionais:
                    glpi.atualizar("Ticket", ticket_id, campos_operacionais)
                if categoria_id:
                    glpi.atualizar("Ticket", ticket_id, {
                        "itilcategories_id": int(categoria_id),
                    })
                if tecnico_id:
                    try:
                        glpi.criar("Ticket_User", {
                            "tickets_id": ticket_id,
                            "users_id": int(tecnico_id),
                            "type": 2,
                        })
                    except Exception as erro_tecnico:
                        registrar_pendencia(
                            pendencias_path,
                            numero,
                            "Técnico",
                            str(erro_tecnico),
                        )
            else:
                ticket_id = glpi.criar("Ticket", ticket)
                if tecnico_id:
                    # "_users_id_assign" na criação às vezes não vincula de
                    # verdade; confere e só cria o vínculo se faltar.
                    try:
                        vinculos_tecnico = glpi.requisicao("GET", f"Ticket/{ticket_id}/Ticket_User?range=0-20")
                        vinculos_tecnico = vinculos_tecnico if isinstance(vinculos_tecnico, list) else []
                    except Exception:
                        vinculos_tecnico = []
                    ja_vinculado = any(
                        int(v.get("users_id") or 0) == tecnico_id and int(v.get("type") or 0) == 2
                        for v in vinculos_tecnico
                    )
                    if not ja_vinculado:
                        try:
                            glpi.criar("Ticket_User", {
                                "tickets_id": ticket_id,
                                "users_id": tecnico_id,
                                "type": 2,
                            })
                        except Exception as erro_tecnico:
                            registrar_pendencia(pendencias_path, numero, "Técnico", str(erro_tecnico))

            # O GLPI bloqueia novos vínculos de documentos em chamados fechados.
            # Por isso, todos os anexos são vinculados antes da solução/fechamento.
            for nome_anexo, url_anexo in links_anexos.get(int(numero.split("-")[1]), []):
                try:
                    caminho_local = localizar_anexo_local(nome_anexo, pasta_anexos_locais)
                    caminho_anexo = caminho_local or baixar_google_drive(nome_anexo, url_anexo, pasta_temporaria)
                    documento_id = glpi.enviar_documento(caminho_anexo)
                    glpi.criar("Document_Item", {
                        "documents_id": documento_id,
                        "items_id": ticket_id,
                        "itemtype": "Ticket",
                    })
                    if caminho_anexo.parent == pasta_temporaria:
                        caminho_anexo.unlink(missing_ok=True)
                except Exception as erro_anexo:
                    registrar_pendencia(pendencias_path, numero, "Anexo", f"{nome_anexo} - {erro_anexo}")

            status_final = STATUS_GLPI.get(status_original, 1)

            solucao = texto(linha.get("Desdobramento"))
            if status_final == 6 and not solucao:
                solucao = "Chamado concluído no controle antigo; a solução não foi registrada no Excel."

            if status_final == 6 and solucao:
                glpi.criar("ITILSolution", {
                    "itemtype": "Ticket",
                    "items_id": ticket_id,
                    "content": html.escape(solucao).replace("\n", "<br>"),
                })

            if data_final and status_final == 6:
                glpi.atualizar("Ticket", ticket_id, {"solvedate": data_final})
            glpi.atualizar("Ticket", ticket_id, {"status": status_final})
            if data_final and status_original == "finalizado":
                registrar_fechamento_sql(sql_fechamentos_path, ticket_id, data_final)

            if acao == "ATUALIZAR":
                estado["importados"][numero]["reparado_em"] = datetime.now().isoformat(timespec="seconds")
                estado["importados"][numero]["status_sincronizado_excel"] = status_final
            elif acao == "RECRIAR":
                agora = datetime.now().isoformat(timespec="seconds")
                estado["importados"][numero] = {
                    "ticket_id": ticket_id,
                    "importado_em": agora,
                    "reparado_em": agora,
                    "substituiu_ticket_ausente": int(ticket_existente),
                    "status_sincronizado_excel": status_final,
                }
            else:
                estado["importados"][numero] = {
                    "ticket_id": ticket_id,
                    "importado_em": datetime.now().isoformat(timespec="seconds"),
                    "status_sincronizado_excel": status_final,
                }
            salvar_estado(estado_path, estado)
            verbo = {
                "ATUALIZAR": "atualizado",
                "RECRIAR": "recriado após ID ausente",
            }.get(acao, "criado")
            print(f"OK: {numero} {verbo} como GLPI #{ticket_id}")
            time.sleep(float(config.get("interval_seconds", 0.3)))
    finally:
        if glpi:
            glpi.finalizar()


if __name__ == "__main__":
    try:
        main()
    except urllib.error.HTTPError as erro:
        corpo = erro.read().decode("utf-8", errors="replace")[:1000]
        print(f"Erro da API: HTTP {erro.code} - {erro.reason}\n{corpo}", file=sys.stderr)
        sys.exit(2)
    except Exception as erro:
        print(f"Erro: {erro}", file=sys.stderr)
        sys.exit(1)
