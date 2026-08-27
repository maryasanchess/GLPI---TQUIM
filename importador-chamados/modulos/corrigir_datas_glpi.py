#!/usr/bin/env python3
"""Compara datas da planilha com GLPI e gera SQL seguro. Nao altera o GLPI."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import ssl
import sys
import zipfile
from collections import Counter
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Dependencias ausentes. Execute o arquivo EXECUTAR_GERACAO_DATAS.bat.")
    raise


FINALIZADO_PLANILHA = "finalizado"
STATUS_GLPI_FINALIZADO = {5, 6}  # solucionado e fechado
COL_ABERTURA = "Carimbo de data/hora"
COL_INICIO_ESCRITO = "Quando o problema iniciou?"
COL_STATUS = "Status"
COL_FINALIZACAO = "Data da Finalização"
COL_NOME = "Nome do Solicitante"
COL_ASSUNTO = "Assunto"
DATE_FIELDS = ("date", "date_creation", "solvedate", "closedate")


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def parse_glpi_datetime(value: Any) -> datetime | None:
    if not value or str(value).strip() in {"", "0", "None"}:
        return None
    text = str(value).strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            pass
    return None


def deterministic_time(key: str) -> time:
    digest = hashlib.sha256(key.encode("utf-8")).digest()
    seconds = int.from_bytes(digest[:4], "big") % (9 * 60 * 60)
    return (datetime.combine(date.today(), time(8, 0)) + timedelta(seconds=seconds)).time()


def clean_year(year: int, reference_year: int | None) -> int:
    if 1900 <= year <= 2100:
        return year
    if reference_year and year % 100 == reference_year % 100:
        return reference_year
    return reference_year or year


def parse_flexible_datetime(value: Any, reference: datetime | None = None) -> datetime | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, time(0, 0))
    if isinstance(value, time):
        return datetime.combine((reference or datetime.now()).date(), value).replace(microsecond=0)
    if isinstance(value, timedelta):
        if value.days > 1000:
            return (datetime(1899, 12, 30) + value).replace(microsecond=0)
        return None
    if isinstance(value, (int, float)):
        if value == 0:
            return None
        if 1 <= float(value) < 2958466:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).replace(microsecond=0)
        return None

    text = str(value).strip()
    if not text or text in {"-", "/", "0"}:
        return None
    text = text.replace("//", "/").replace("::", ":").replace("/.", ":")
    text = re.sub(r"\s*-\s*", " ", text)
    text = re.sub(r"(?<=\d)(?=\d{2}:\d{2}$)", " ", text)
    text = re.sub(r":+$", "", text)

    iso = re.search(r"\b(20\d{2})-(\d{1,2})-(\d{1,2})(?:[ T]+(\d{1,2})[:.](\d{2})(?::(\d{2}))?)?", text)
    if iso:
        y, m, d = map(int, iso.group(1, 2, 3))
        hh, mm, ss = (int(x or 0) for x in iso.group(4, 5, 6))
        try:
            return datetime(y, m, d, hh, mm, ss)
        except ValueError:
            return None

    br = re.search(r"\b(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?\b", text)
    clock = re.search(r"\b([01]?\d|2[0-3])[:.](\d{2})(?::(\d{2}))?\b", text)
    if br:
        d, m = int(br.group(1)), int(br.group(2))
        raw_year = br.group(3)
        if raw_year:
            y = int(raw_year)
            if y < 100:
                y += 2000
            y = clean_year(y, reference.year if reference else None)
        elif reference:
            y = reference.year
        else:
            return None
        hh, mm, ss = (int(clock.group(i) or 0) if clock else 0 for i in (1, 2, 3))
        try:
            return datetime(y, m, d, hh, mm, ss)
        except ValueError:
            return None

    if clock and reference:
        return datetime.combine(reference.date(), time(int(clock.group(1)), int(clock.group(2)), int(clock.group(3) or 0)))
    return None


def raw_numeric_cells(xlsx_path: Path, column_letter: str) -> dict[int, float]:
    result: dict[int, float] = {}
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        with zipfile.ZipFile(xlsx_path) as archive:
            root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        for cell in root.findall(".//m:c", ns):
            ref = cell.attrib.get("r", "")
            if not ref.startswith(column_letter):
                continue
            value = cell.find("m:v", ns)
            if value is None:
                continue
            try:
                result[int(re.sub(r"\D", "", ref))] = float(value.text)
            except (TypeError, ValueError):
                pass
    except Exception:
        pass
    return result


def read_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {COL_ABERTURA, COL_INICIO_ESCRITO, COL_STATUS, COL_FINALIZACAO}
    missing = required - set(headers)
    if missing:
        raise RuntimeError("Colunas obrigatorias ausentes: " + ", ".join(sorted(missing)))
    raw_final = raw_numeric_cells(xlsx_path, get_column_letter(headers.index(COL_FINALIZACAO) + 1))
    rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = dict(zip(headers, values))
        item["_excel_row"] = excel_row
        item["_raw_final"] = raw_final.get(excel_row)
        rows.append(item)
    meaningful = (COL_ABERTURA, COL_NOME, COL_ASSUNTO, COL_STATUS, COL_FINALIZACAO)
    while rows and not any(norm_text(rows[-1].get(column)) for column in meaningful):
        rows.pop()
    return rows


def pick_opening(row: dict[str, Any], key: str) -> tuple[datetime | None, str]:
    opening = parse_flexible_datetime(row.get(COL_ABERTURA))
    if opening:
        return opening.replace(microsecond=0), "abertura da planilha"
    written = parse_flexible_datetime(row.get(COL_INICIO_ESCRITO))
    if written:
        return written.replace(microsecond=0), "abertura derivada da data escrita"
    final = parse_flexible_datetime(row.get(COL_FINALIZACAO))
    if final:
        base = final - timedelta(days=3)
        return datetime.combine(base.date(), deterministic_time(key + ":open")), "abertura derivada da finalizacao menos 3 dias"
    return None, "abertura indefinida"


def pick_final(row: dict[str, Any], opening: datetime, key: str, now: datetime) -> tuple[datetime, str]:
    value = row.get(COL_FINALIZACAO)
    candidate = parse_flexible_datetime(value, opening)
    if candidate is None and isinstance(value, timedelta) and row.get("_raw_final"):
        raw = row["_raw_final"]
        if 1 <= raw < 2958466:
            candidate = datetime(1899, 12, 30) + timedelta(days=raw)
    if candidate:
        candidate = candidate.replace(microsecond=0)
        if candidate >= opening and candidate <= now + timedelta(days=1):
            return candidate, "finalizacao valida da planilha"

    written = parse_flexible_datetime(row.get(COL_INICIO_ESCRITO), opening)
    if written and written.date() > opening.date() and written <= now:
        return datetime.combine(written.date(), deterministic_time(key + ":written")), "finalizacao baseada na data escrita"

    fallback_day = opening.date() + timedelta(days=3)
    fallback = datetime.combine(fallback_day, deterministic_time(key + ":final"))
    if fallback > now:
        fallback = datetime.combine(now.date(), deterministic_time(key + ":current"))
        if fallback < opening:
            fallback = opening + timedelta(minutes=30)
    return fallback.replace(microsecond=0), "finalizacao estimada: abertura + 3 dias, horario 08h-17h"


class GlpiClient:
    def __init__(self, config: dict[str, Any]):
        base = config.get("glpi_url") or config.get("api_url") or config.get("url_glpi") or config.get("base_url") or config.get("url")
        if not base:
            raise RuntimeError("Informe glpi_url (ou api_url) no config_glpi.json")
        base = str(base).rstrip("/")
        self.api = base if base.endswith("apirest.php") else base + "/apirest.php"
        self.app_token = config.get("app_token") or config.get("apptoken")
        self.user_token = config.get("user_token") or config.get("usertoken")
        if not self.app_token or not self.user_token:
            raise RuntimeError("Informe app_token e user_token no config_glpi.json")
        self.verify = bool(config.get("verificar_ssl", True))
        self.ssl_context = None if self.verify else ssl._create_unverified_context()
        self.timeout = int(config.get("timeout_segundos", 60))
        self.session_token: str | None = None

    def get_json(self, endpoint: str, headers: dict[str, str], params: dict[str, str] | None = None) -> Any:
        url = self.api + endpoint
        if params:
            url += "?" + urlencode(params)
        request = Request(url, headers=headers, method="GET")
        with urlopen(request, timeout=self.timeout, context=self.ssl_context) as response:
            return json.loads(response.read().decode("utf-8-sig"))

    def init(self) -> None:
        data = self.get_json(
            "/initSession",
            {"Authorization": f"user_token {self.user_token}", "App-Token": str(self.app_token)},
        )
        self.session_token = data["session_token"]

    def headers(self) -> dict[str, str]:
        return {"Session-Token": str(self.session_token), "App-Token": str(self.app_token)}

    def get_all_tickets(self) -> list[dict[str, Any]]:
        tickets: list[dict[str, Any]] = []
        start, batch = 0, 1000
        while True:
            try:
                items = self.get_json(
                    "/Ticket",
                    self.headers(),
                    {"range": f"{start}-{start + batch - 1}", "get_hateoas": "false"},
                )
            except HTTPError as exc:
                if exc.code == 416:
                    break
                raise
            if not isinstance(items, list):
                raise RuntimeError("Resposta inesperada ao consultar chamados no GLPI")
            tickets.extend(items)
            if len(items) < batch:
                break
            start += batch
        return tickets

    def close(self) -> None:
        if not self.session_token:
            return
        try:
            self.get_json("/killSession", self.headers())
        except Exception:
            pass


MARKER_RE = re.compile(r"\bEXCEL\s*[-#:]?\s*(\d+)\b", re.IGNORECASE)


def ticket_marker(ticket: dict[str, Any]) -> int | None:
    text = " ".join(str(ticket.get(k) or "") for k in ("name", "content"))
    match = MARKER_RE.search(text)
    return int(match.group(1)) if match else None


def infer_offset(tickets: list[dict[str, Any]], rows: list[dict[str, Any]]) -> tuple[int, list[tuple[int, int, int]]]:
    scores: list[tuple[int, int, int]] = []
    for offset in range(-10, 11):
        exact = near = 0
        for ticket in tickets:
            marker = ticket_marker(ticket)
            if marker is None:
                continue
            idx = marker - 1 + offset
            if not (0 <= idx < len(rows)):
                continue
            source = parse_flexible_datetime(rows[idx].get(COL_ABERTURA))
            current = parse_glpi_datetime(ticket.get("date"))
            if not source or not current:
                continue
            delta = abs((source.date() - current.date()).days)
            if delta == 0:
                exact += 1
            if delta <= 3:
                near += 1
        scores.append((exact, near, offset))
    scores.sort(reverse=True)
    best = scores[0]
    if best[0] < 10:
        raise RuntimeError("Nao foi possivel confirmar automaticamente a correspondencia EXCEL-N com as linhas da planilha")
    return best[2], scores[:5]


def differs(a: datetime | None, b: datetime | None, tolerance_seconds: int) -> bool:
    if a is None or b is None:
        return a != b
    return abs((a - b).total_seconds()) > tolerance_seconds


def sql_literal(value: datetime | None) -> str:
    return "NULL" if value is None else "'" + value.strftime("%Y-%m-%d %H:%M:%S") + "'"


def old_guard(field: str, value: datetime | None) -> str:
    return f"`{field}` <=> {sql_literal(value)}"


def build_update(ticket_id: int, old: dict[str, datetime | None], new: dict[str, datetime | None], changed: list[str]) -> str:
    assignments = ", ".join(f"`{field}` = {sql_literal(new[field])}" for field in changed)
    guards = " AND ".join(old_guard(field, old[field]) for field in changed)
    return f"UPDATE `glpi_tickets` SET {assignments} WHERE `id` = {ticket_id} AND {guards};"


def save_report(path: Path, records: list[dict[str, Any]], summary: Counter, metadata: list[tuple[str, Any]]) -> None:
    wb = load_workbook(Path(__file__).with_name("modelo_relatorio.xlsx")) if Path(__file__).with_name("modelo_relatorio.xlsx").exists() else None
    if wb is None:
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "Resumo"
    ws = wb["Resumo"]
    ws.delete_rows(1, ws.max_row)
    ws.append(["CORRECAO DE DATAS GLPI - SOMENTE DIVERGENCIAS"])
    ws.append([])
    for key, value in metadata:
        ws.append([key, value])
    ws.append([])
    ws.append(["Resultado", "Quantidade"])
    for key, value in sorted(summary.items()):
        ws.append([key, value])

    if "Plano" in wb.sheetnames:
        del wb["Plano"]
    plan = wb.create_sheet("Plano")
    headers = [
        "Resultado", "GLPI ID", "Referencia", "Linha Excel", "Solicitante", "Assunto", "Status GLPI",
        "Abertura GLPI", "Abertura alvo", "Criacao GLPI", "Criacao alvo",
        "Solucao GLPI", "Solucao alvo", "Fechamento GLPI", "Fechamento alvo",
        "Campos divergentes", "Regra da abertura", "Regra da finalizacao",
    ]
    plan.append(headers)
    for record in records:
        plan.append([record.get(h, "") for h in headers])

    for sheet in (ws, plan):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            values = [str(sheet.cell(r, col).value or "") for r in range(1, min(sheet.max_row, 200) + 1)]
            width = min(max(max(map(len, values), default=8) + 2, 10), 42)
            sheet.column_dimensions[get_column_letter(col)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws.merge_cells("A1:B1")
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Gera SQL para corrigir apenas datas divergentes no GLPI")
    parser.add_argument("--config", default="config_glpi.json")
    parser.add_argument("--planilha", default=None)
    args = parser.parse_args()

    base = Path.cwd()
    config_path = (base / args.config).resolve()
    if not config_path.exists():
        raise RuntimeError(f"Arquivo de configuracao nao encontrado: {config_path.name}")
    config = json.loads(config_path.read_text(encoding="utf-8-sig"))
    xlsx_path = (base / (args.planilha or config.get("planilha") or "planilha_atualizada.xlsx")).resolve()
    if not xlsx_path.exists():
        raise RuntimeError(f"Planilha nao encontrada: {xlsx_path.name}")

    tolerance = int(config.get("tolerancia_segundos", 60))
    fix_creation = bool(config.get("corrigir_data_criacao", True))
    process_nonfinal = bool(config.get("processar_nao_finalizados", False))
    now = datetime.now().replace(microsecond=0)
    rows = read_rows(xlsx_path)

    client = GlpiClient(config)
    try:
        client.init()
        tickets = client.get_all_tickets()
    finally:
        client.close()

    print(f"Consultados {len(tickets)} chamados no GLPI.")
    offset, offset_scores = infer_offset(tickets, rows)
    print(f"Correspondencia confirmada: indice EXCEL-N + deslocamento {offset}.")

    records: list[dict[str, Any]] = []
    summary: Counter = Counter()
    updates: list[str] = []
    update_ids: list[int] = []

    for ticket in tickets:
        marker = ticket_marker(ticket)
        if marker is None:
            continue
        ticket_id = int(ticket["id"])
        idx = marker - 1 + offset
        reference = f"EXCEL-{marker}"
        base_record = {
            "GLPI ID": ticket_id,
            "Referencia": reference,
            "Status GLPI": ticket.get("status"),
        }
        if not (0 <= idx < len(rows)):
            result = "IGNORADO - LINHA NAO LOCALIZADA"
            summary[result] += 1
            records.append({"Resultado": result, **base_record})
            continue
        row = rows[idx]
        base_record.update({
            "Linha Excel": row["_excel_row"],
            "Solicitante": row.get(COL_NOME),
            "Assunto": row.get(COL_ASSUNTO),
        })
        if norm_text(row.get(COL_NOME)) in {"", "ignorar", "nan"}:
            result = "IGNORADO - LINHA VAZIA/IGNORAR"
            summary[result] += 1
            records.append({"Resultado": result, **base_record})
            continue

        excel_finalized = norm_text(row.get(COL_STATUS)) == FINALIZADO_PLANILHA
        glpi_status = int(ticket.get("status") or 0)
        if not process_nonfinal and (not excel_finalized or glpi_status not in STATUS_GLPI_FINALIZADO):
            result = "IGNORADO - CHAMADO NAO FINALIZADO"
            summary[result] += 1
            records.append({"Resultado": result, **base_record})
            continue

        opening, opening_rule = pick_opening(row, reference)
        if opening is None:
            result = "IGNORADO - ABERTURA INDEFINIDA"
            summary[result] += 1
            records.append({"Resultado": result, **base_record, "Regra da abertura": opening_rule})
            continue
        final, final_rule = pick_final(row, opening, reference, now)
        old = {field: parse_glpi_datetime(ticket.get(field)) for field in DATE_FIELDS}
        new = dict(old)
        new["date"] = opening
        if fix_creation:
            new["date_creation"] = opening
        if excel_finalized and glpi_status in STATUS_GLPI_FINALIZADO:
            new["solvedate"] = final
            if glpi_status == 6:
                new["closedate"] = final

        changed = [field for field in DATE_FIELDS if differs(old[field], new[field], tolerance)]
        result = "PRONTO PARA CORRIGIR" if changed else "SEM ALTERACAO"
        summary[result] += 1
        if changed:
            updates.append(build_update(ticket_id, old, new, changed))
            update_ids.append(ticket_id)
        records.append({
            "Resultado": result,
            **base_record,
            "Abertura GLPI": old["date"], "Abertura alvo": new["date"],
            "Criacao GLPI": old["date_creation"], "Criacao alvo": new["date_creation"],
            "Solucao GLPI": old["solvedate"], "Solucao alvo": new["solvedate"],
            "Fechamento GLPI": old["closedate"], "Fechamento alvo": new["closedate"],
            "Campos divergentes": ", ".join(changed),
            "Regra da abertura": opening_rule,
            "Regra da finalizacao": final_rule,
        })

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = base / f"resultado_datas_{stamp}"
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"PLANO_CORRECOES_DATAS_{stamp}.xlsx"
    metadata = [
        ("Planilha", xlsx_path.name),
        ("Chamados consultados", len(tickets)),
        ("Deslocamento EXCEL-N", offset),
        ("Melhores pontuacoes", str(offset_scores)),
        ("Data de geracao", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Modo", "SOMENTE GERACAO DE SQL - NENHUMA ALTERACAO NO GLPI"),
    ]
    save_report(report_path, records, summary, metadata)

    ids_sql = ", ".join(map(str, sorted(set(update_ids)))) or "0"
    test_path = output_dir / f"SQL_1_TESTE_ROLLBACK_{stamp}.sql"
    apply_path = output_dir / f"SQL_2_APLICAR_COMMIT_{stamp}.sql"
    test_lines = [
        "-- TESTE: TODAS AS ALTERACOES SERAO DESFEITAS COM ROLLBACK",
        "START TRANSACTION;",
        f"SELECT id, date, date_creation, solvedate, closedate FROM glpi_tickets WHERE id IN ({ids_sql}) ORDER BY id;",
        *updates,
        f"SELECT id, date, date_creation, solvedate, closedate FROM glpi_tickets WHERE id IN ({ids_sql}) ORDER BY id;",
        "ROLLBACK;",
    ]
    backup_table = f"backup_glpi_tickets_datas_{stamp}"
    apply_lines = [
        "-- APLICACAO DEFINITIVA: CONFIRA PRIMEIRO O RELATORIO E O SQL DE TESTE",
        f"CREATE TABLE IF NOT EXISTS `{backup_table}` LIKE `glpi_tickets`;",
        f"INSERT IGNORE INTO `{backup_table}` SELECT * FROM `glpi_tickets` WHERE `id` IN ({ids_sql});",
        "START TRANSACTION;",
        *updates,
        "COMMIT;",
        f"SELECT id, date, date_creation, solvedate, closedate FROM glpi_tickets WHERE id IN ({ids_sql}) ORDER BY id;",
    ]
    test_path.write_text("\n".join(test_lines) + "\n", encoding="utf-8")
    apply_path.write_text("\n".join(apply_lines) + "\n", encoding="utf-8")

    print("\nRESUMO DO PLANO")
    for key, value in sorted(summary.items()):
        print(f"  {key}: {value}")
    print(f"\nArquivos gerados em: {output_dir}")
    print("Nenhuma data foi alterada no GLPI.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"\nERRO: {exc}")
        raise SystemExit(1)
