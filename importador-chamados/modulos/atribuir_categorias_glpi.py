#!/usr/bin/env python3
"""Atribui categorias somente a chamados do GLPI que ainda estão sem categoria.

O arquivo Excel deve possuir as colunas ID, Título e Categoria. Quando Categoria
estiver vazia, o script sugere uma categoria com base no título. A execução real
só ocorre com --executar; sem essa opção, o programa apenas simula.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


CATEGORIA_A_CLASSIFICAR = "A classificar"


def texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def normalizar(valor) -> str:
    valor = unicodedata.normalize("NFKD", texto(valor))
    valor = "".join(ch for ch in valor if not unicodedata.combining(ch))
    valor = valor.casefold().replace("→", ">")
    valor = re.sub(r"\s*>\s*", " > ", valor)
    valor = re.sub(r"\s+", " ", valor)
    return valor.strip()


def contem(texto_normalizado: str, *termos: str) -> bool:
    return any(normalizar(termo) in texto_normalizado for termo in termos)


def id_inteiro(valor) -> int | None:
    digitos = re.sub(r"\D", "", texto(valor))
    return int(digitos) if digitos else None


def titulo_limpo(titulo: str) -> str:
    titulo = re.sub(r"^\s*\[MIGRAÇÃO\s+EXCEL-\d+\]\s*", "", texto(titulo), flags=re.I)
    titulo = re.sub(r"^(Outros|Zion)\s*-\s*", "", titulo, flags=re.I)
    return titulo.strip()


def categoria_sugerida(titulo: str, ticket_id: int | None = None) -> str:
    original = texto(titulo)
    t = normalizar(original)
    assunto = normalizar(titulo_limpo(original))

    # Registros de teste, vazios ou sem informação suficiente.
    if ticket_id in {1, 2, 10, 19, 20, 24}:
        return CATEGORIA_A_CLASSIFICAR
    if not assunto or assunto in {"teste", "bom dia", "bom dia!", "boa tarde", "boa tarde!", "sem descricao"}:
        return CATEGORIA_A_CLASSIFICAR
    if contem(assunto, "abertura de chamado - teste", "teste grupo do usuario", "aeadfsdfasdf", "vldihne"):
        return CATEGORIA_A_CLASSIFICAR

    # Zion e WMS devem ser avaliados antes das regras genéricas de acesso.
    if "zion -" in t or contem(t, "zion", "skyone", "skayon"):
        if contem(assunto, "wms"):
            return "Zion > WMS"
        if contem(assunto, "nota fiscal", "integracao", "integrado", "emitir notas"):
            return "Zion > Integração e Nota Fiscal"
        if contem(assunto, "usuario bloqueado", "login", "senha", "acesso"):
            return "Zion > Acesso e Login"
        return "Zion > Instabilidade ou Indisponibilidade"

    # Certificado digital aparece com grande frequência no histórico.
    if contem(assunto, "certificado", "e-cnpj", "ecnpj", "assinador", "assinatura digital"):
        if contem(assunto, "assinador", "assinatura digital"):
            return "Certificado Digital > Assinador e Assinatura Digital"
        if contem(assunto, "atualizacao", "atualizar", "renovacao", "renovar"):
            return "Certificado Digital > Atualização e Renovação"
        if contem(assunto, "nao funciona", "nao esta funcionando", "erro", "verificar", "bloqueado"):
            return "Certificado Digital > Erro ou Certificado não Funciona"
        return "Certificado Digital > Instalação"

    # Backup, perda e recuperação precisam vir antes de diretórios e Excel.
    if contem(assunto, "recuperar", "recuperacao", "restauracao", "backup", "backcap"):
        if contem(assunto, "recuperar", "recuperacao", "restauracao", "excluid", "apagada", "deletad", "sumiu"):
            return "Backup e Arquivos > Recuperação de Arquivos"
        return "Backup e Arquivos > Realização de Backup"
    if contem(assunto, "arquivo excluido", "arquivos excluidos", "pasta excluida", "planilha excluida", "apagada por engano", "sumiu do diretorio"):
        return "Backup e Arquivos > Arquivo ou Pasta Excluída"
    if contem(assunto, "arquivo corrompido", "planilha dando erro ao abrir"):
        return "Backup e Arquivos > Arquivo Corrompido ou não Abre"

    # Rodopar e documentos operacionais.
    if contem(assunto, "rodopar"):
        if contem(assunto, "faturamento"):
            return "Rodopar > Faturamento"
        if contem(assunto, "programacao", "grade", "carga"):
            return "Rodopar > Programação"
        if contem(assunto, "fiscal", "cte", "ct-e", "mdf", "nota fiscal", "xml"):
            return "Rodopar > Fiscal"
        if contem(assunto, "frota", "sinistro", "colisao"):
            return "Rodopar > Frota"
        if contem(assunto, "manutencao", "o.s", "os "):
            return "Rodopar > Manutenção"
        return "Acessos e Permissões > Sistemas"
    if contem(assunto, "cte ", "ct-e", "mdf ", "mdf-"):
        return "Rodopar > Fiscal"
    if contem(assunto, "divergencia de faturamento - sped"):
        return "Rodopar > Faturamento"
    if contem(assunto, "recebimento de nota fiscal"):
        return "Rodopar > Fiscal"
    if contem(assunto, "grade de programacao", "terminal de carregamento", "abertura de o.s", "os nº"):
        return "Rodopar > Programação"

    # Impressão e digitalização.
    if contem(assunto, "scanner", "scaner"):
        return "Impressora > Scanner"
    if contem(assunto, "impressora"):
        if contem(assunto, "instalar", "configurar", "configurad", "habilitar"):
            return "Impressora > Instalação e Configuração"
        if contem(assunto, "servidor", "conecta"):
            return "Impressora > Conexão com Servidor"
        return "Impressora > Erro de Impressão"

    # Câmeras: diferenciar pedido de acesso de defeito/instalação física.
    if contem(assunto, "camera", "cameras", "dvr"):
        if contem(assunto, "acesso", "liberar", "inclusao", "visualizar"):
            return "Acessos e Permissões > Câmeras"
        return "Hardware (Equipamentos) > Câmeras e DVR"

    # Hardware e periféricos.
    if contem(assunto, "nobreak", "no break", "estabilizador", "surto de tensao", "transformador"):
        return "Hardware (Equipamentos) > Nobreak e Estabilizador"
    if contem(assunto, "monitor", "segunda tela", "duas telas", "tela da tv", "televisao", "tvs ", "tv ", "hdmi"):
        return "Hardware (Equipamentos) > Monitor e TV"
    if contem(assunto, "mouse", "teclado", "fone de ouvido", "mouse pad", "pilhas aaa", "caixinha de som", "porta usb"):
        return "Hardware (Equipamentos) > Mouse, Teclado e Periféricos"
    if contem(assunto, "linha telefonica", "ponto telefonico", "pontos telefonicos", "telefone"):
        return "Hardware (Equipamentos) > Telefonia"
    if contem(assunto, "movimentacao de equipamentos", "mudanca de lugar", "realocar", "alteracao do computador", "transferir os dados do meu computador"):
        return "Hardware (Equipamentos) > Instalação e Movimentação de Equipamentos"
    if contem(assunto, "solicito um novo", "solicitacao de troca de notebook", "precisando de 1 monitor", "trocar o monitor"):
        return "Hardware (Equipamentos) > Solicitação de Equipamentos"
    if contem(assunto, "computador", "notebook", "notbook", "memoria"):
        if contem(assunto, "configurar", "configuracao", "computador novo", "instalacoes dos computadores"):
            return "Hardware (Equipamentos) > Instalação e Movimentação de Equipamentos"
        return "Hardware (Equipamentos) > Computador e Notebook"

    # E-mail antes das regras genéricas de usuário e software.
    if contem(assunto, "grupo rh.vagas", "grupo de e-mail", "grupo do e-mail"):
        return "E-mail > Grupos de E-mail"
    if contem(assunto, "assinatura de e-mail", "e-mail sem assinatura", "assinatura atualizada"):
        return "E-mail > Assinatura de E-mail"
    if contem(assunto, "e-mail", "email", "emil"):
        if contem(assunto, "criar", "criacao", "configurar", "configuracao", "transferir", "desvincular"):
            return "E-mail > Criação e Configuração"
        return "E-mail > Envio e Recebimento"

    # Acessos, contas, diretórios, sites e FTP.
    if contem(assunto, "ftp"):
        return "Acessos e Permissões > FTP"
    if contem(assunto, "diretorio", "diretorios", "pasta ", "pastas "):
        if contem(assunto, "erro", "nao aparece", "sumiu", "nao consigo pesquisar"):
            return "Backup e Arquivos > Arquivo ou Pasta Excluída"
        return "Acessos e Permissões > Diretórios e Pastas"
    if contem(assunto, "mapear diretorio", "mapear direorio"):
        return "Acessos e Permissões > Diretórios e Pastas"
    if contem(assunto, "site bloqueado", "sites bloqueados", "liberacao do site", "liberar o site", "recaptcha"):
        return "Acessos e Permissões > Sites"
    if contem(assunto, "acessar alguns sites", "sites de empresa"):
        return "Acessos e Permissões > Sites"
    if contem(assunto, "senha", "login", "usuario", "usuário", "criar acesso", "criacao de acesso"):
        return "Acessos e Permissões > Usuários, Logins e Senhas"

    # Aplicativos e documentos.
    if contem(assunto, "power bi"):
        return "Software > Power BI"
    if contem(assunto, "excel", "planilha", "office"):
        return "Software > Microsoft Office e Excel"
    if contem(assunto, "pdf", "danfe view", "xml", "visualizador de xml"):
        return "Software > PDF, XML e DANFE View"
    if contem(assunto, "windows"):
        return "Software > Windows"
    if contem(assunto, "limpeza de cache"):
        return "Software > Windows"
    if contem(assunto, "instalacao de java", "instalar java"):
        return "Software > Instalação e Atualização"
    if contem(assunto, "whatsapp", "teams"):
        return "Software > WhatsApp e Teams"
    if contem(assunto, "anydesk"):
        return "Software > AnyDesk e Acesso Remoto"

    # Sistemas corporativos recorrentes.
    if contem(assunto, "datapar"):
        return "Sistemas Corporativos > Datapar"
    if contem(assunto, "onixsat", "onix", "omisat"):
        return "Sistemas Corporativos > Rastreamento > OnixSat e Omisat"
    if contem(assunto, "autotrac"):
        return "Sistemas Corporativos > Rastreamento > Autotrac"
    if contem(assunto, "transport", "e-fleet", "efleet", "truckscontrol"):
        return "Sistemas Corporativos > Rastreamento > Transport e E-Fleet"
    if contem(assunto, "webcol", "tracking evonik"):
        return "Sistemas Corporativos > Rastreamento > Webcol"
    if contem(assunto, "itau", "banco do brasil", "sispag", "aplicativo do banco"):
        return "Sistemas Corporativos > Aplicativos Bancários"
    if contem(assunto, "sefaz", "nfe fazenda", "nf-e", "sintegra", "nota fiscal eletronica"):
        return "Sistemas Corporativos > Sistemas Fiscais"
    if contem(assunto, "ponto eletronico", "pwponto", "ponto sem comunicacao", "baixo o ponto"):
        return "Sistemas Corporativos > Ponto Eletrônico"
    if contem(assunto, "sap"):
        return "Sistemas Corporativos > SAP"
    if contem(assunto, "dropbox"):
        return "Software > Instalação e Atualização"
    if contem(assunto, "aplicativo", "plataforma", "sistema ", "picture", "mednet", "santos brasil"):
        return "Sistemas Corporativos > Outros Sistemas"
    if contem(assunto, "baixar umas imagens de uns exames", "abertura de bis"):
        return "Sistemas Corporativos > Outros Sistemas"

    # Rede e infraestrutura.
    if contem(assunto, "servidor"):
        return "Rede/Internet > Servidor"
    if contem(assunto, "wi-fi", "wifi", "internet", "conexao", "rede "):
        return "Rede/Internet > Internet e Wi-Fi"
    if contem(assunto, "switch", "ponto de rede"):
        return "Rede/Internet > Ponto de Rede e Switch"
    if contem(assunto, "lento", "lentidao", "instabilidade"):
        return "Rede/Internet > Lentidão e Instabilidade"

    # Solicitações de melhoria, integração e alteração.
    if contem(assunto, "integracao", "integralizacao", "envio automatico", "automacao"):
        return "Desenvolvimento e Melhorias > Automação e Integração"
    if contem(assunto, "criar a saida na torre", "nova funcionalidade"):
        return "Desenvolvimento e Melhorias > Nova Funcionalidade"
    if contem(assunto, "inserir", "inclusao", "alteracao", "ajuste", "exclusao de uma macro", "adicionar no campo"):
        return "Desenvolvimento e Melhorias > Alteração de Sistema"
    if contem(assunto, "relatorio", "formula"):
        return "Desenvolvimento e Melhorias > Relatórios e Planilhas"
    if contem(assunto, "nao estou conseguindo salvar documentos"):
        return "Backup e Arquivos > Arquivo Corrompido ou não Abre"
    if contem(assunto, "funcionario", "jovem aprendiz"):
        return "Acessos e Permissões > Usuários, Logins e Senhas"

    return CATEGORIA_A_CLASSIFICAR


def carregar_config(caminho: Path) -> dict:
    with caminho.open("r", encoding="utf-8") as arquivo:
        config = json.load(arquivo)
    obrigatorios = ("api_url", "app_token", "user_token")
    ausentes = [campo for campo in obrigatorios if not texto(config.get(campo)) or "COLE_" in texto(config.get(campo))]
    if ausentes:
        raise ValueError("Preencha no config_categorias.json: " + ", ".join(ausentes))
    config["api_url"] = texto(config["api_url"]).rstrip("/")
    return config


class Glpi:
    def __init__(self, config: dict):
        self.config = config
        self.headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "App-Token": config["app_token"],
            "Authorization": f"user_token {config['user_token']}",
        }
        self.session_token = None

    def requisicao(self, metodo: str, endpoint: str, dados=None, timeout: int = 45):
        corpo = json.dumps(dados, ensure_ascii=False).encode("utf-8") if dados is not None else None
        requisicao = urllib.request.Request(
            f"{self.config['api_url']}/{endpoint.lstrip('/')}",
            data=corpo,
            headers=self.headers,
            method=metodo,
        )
        try:
            with urllib.request.urlopen(requisicao, timeout=timeout) as resposta:
                conteudo = resposta.read().decode("utf-8")
        except urllib.error.HTTPError as erro:
            detalhe = erro.read().decode("utf-8", errors="replace")[:1500]
            raise RuntimeError(f"Falha na API em {metodo} {endpoint}: HTTP {erro.code} - {detalhe}") from erro
        return json.loads(conteudo) if conteudo else None

    def iniciar(self):
        resposta = self.requisicao("GET", "initSession", timeout=30)
        self.session_token = resposta["session_token"]
        self.headers["Session-Token"] = self.session_token
        perfil_id = self.config.get("profile_id", 4)
        if perfil_id:
            self.requisicao("POST", "changeActiveProfile", {"profiles_id": int(perfil_id)})

    def finalizar(self):
        if self.session_token:
            try:
                self.requisicao("GET", "killSession", timeout=15)
            except Exception:
                pass

    def listar_categorias(self) -> list[dict]:
        retorno = self.requisicao("GET", "ITILCategory?range=0-9999")
        return retorno if isinstance(retorno, list) else []

    def obter_ticket(self, ticket_id: int) -> dict | None:
        try:
            retorno = self.requisicao("GET", f"Ticket/{ticket_id}")
        except RuntimeError as erro:
            if "HTTP 404" in str(erro) or "ITEM_NOT_FOUND" in str(erro):
                return None
            raise
        return retorno if isinstance(retorno, dict) else None

    def atribuir_categoria(self, ticket_id: int, categoria_id: int):
        self.requisicao(
            "PUT",
            f"Ticket/{ticket_id}",
            {"input": {"id": ticket_id, "itilcategories_id": categoria_id}},
        )


def localizar_colunas(planilha) -> dict[str, int]:
    cabecalhos = {normalizar(c.value): c.column for c in planilha[1] if texto(c.value)}
    aliases = {
        "id": ("id", "id glpi", "ticket", "ticket id"),
        "titulo": ("titulo", "título", "nome"),
        "categoria": ("categoria", "categoria sugerida"),
    }
    encontradas = {}
    for destino, opcoes in aliases.items():
        for opcao in opcoes:
            if normalizar(opcao) in cabecalhos:
                encontradas[destino] = cabecalhos[normalizar(opcao)]
                break
    faltantes = [nome for nome in ("id", "titulo") if nome not in encontradas]
    if faltantes:
        raise ValueError("A planilha precisa possuir as colunas ID e Título.")
    return encontradas


def ler_planilha(caminho: Path) -> list[dict]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    planilha = workbook.active
    colunas = localizar_colunas(planilha)
    registros = []
    vistos = set()
    for numero_linha in range(2, planilha.max_row + 1):
        ticket_id = id_inteiro(planilha.cell(numero_linha, colunas["id"]).value)
        titulo = texto(planilha.cell(numero_linha, colunas["titulo"]).value)
        categoria_excel = texto(planilha.cell(numero_linha, colunas["categoria"]).value) if "categoria" in colunas else ""
        if not ticket_id or not titulo:
            continue
        if ticket_id in vistos:
            raise ValueError(f"ID duplicado na planilha: GLPI #{ticket_id}")
        vistos.add(ticket_id)
        categoria = categoria_excel or categoria_sugerida(titulo, ticket_id)
        registros.append({
            "linha": numero_linha,
            "id": ticket_id,
            "titulo": titulo,
            "origem_categoria": "planilha" if categoria_excel else "regra automática",
            "categoria": categoria,
        })
    return registros


def mapa_categorias_glpi(categorias: list[dict], config: dict) -> dict[str, int]:
    mapa = {}
    for item in categorias:
        if not item.get("id"):
            continue
        nome = texto(item.get("completename") or item.get("name"))
        if nome:
            mapa[normalizar(nome)] = int(item["id"])
    for nome, categoria_id in config.get("category_map", {}).items():
        mapa[normalizar(nome)] = int(categoria_id)
    return mapa


def categoria_atual_id(ticket: dict) -> int:
    valor = ticket.get("itilcategories_id")
    if isinstance(valor, dict):
        valor = valor.get("id")
    try:
        return int(valor or 0)
    except (TypeError, ValueError):
        return 0


def salvar_relatorio(caminho: Path, linhas: list[dict]):
    with caminho.open("w", encoding="utf-8-sig", newline="") as arquivo:
        campos = ["ID", "Título", "Categoria", "Origem", "Resultado", "Detalhe"]
        escritor = csv.DictWriter(arquivo, fieldnames=campos, delimiter=";")
        escritor.writeheader()
        escritor.writerows(linhas)


def analisar_local(registros: list[dict], relatorio: Path | None):
    contagem = Counter(item["categoria"] for item in registros)
    print(f"ANÁLISE LOCAL: {len(registros)} chamados encontrados na planilha.")
    for categoria, quantidade in contagem.most_common():
        print(f"{quantidade:3}  {categoria}")
    if relatorio:
        linhas = [{
            "ID": item["id"],
            "Título": item["titulo"],
            "Categoria": item["categoria"],
            "Origem": item["origem_categoria"],
            "Resultado": "SUGESTÃO",
            "Detalhe": "Revisar antes da execução",
        } for item in registros]
        salvar_relatorio(relatorio, linhas)
        print(f"Relatório de revisão criado: {relatorio}")


def main():
    parser = argparse.ArgumentParser(
        description="Atribui categorias apenas aos chamados do GLPI que ainda estão sem categoria."
    )
    parser.add_argument("excel", help="Planilha .xlsx exportada do GLPI")
    parser.add_argument("--config", default="config_categorias.json", help="Arquivo de configuração da API")
    parser.add_argument("--limit", type=int, default=5, help="Máximo de chamados a atualizar; 0 significa todos")
    parser.add_argument("--executar", action="store_true", help="Realiza as alterações; sem isso apenas simula")
    parser.add_argument("--somente-classificar", action="store_true", help="Analisa a planilha sem conectar ao GLPI")
    parser.add_argument("--incluir-a-classificar", action="store_true", help="Permite atribuir a categoria 'A classificar'")
    parser.add_argument("--ticket", type=int, action="append", help="Processa somente o ID informado; pode repetir")
    parser.add_argument("--relatorio", default="resultado_categorias_glpi.csv", help="Arquivo CSV de resultado")
    args = parser.parse_args()

    excel = Path(args.excel).resolve()
    if not excel.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {excel}")
    registros = ler_planilha(excel)
    if args.ticket:
        ids = set(args.ticket)
        registros = [item for item in registros if item["id"] in ids]

    relatorio = Path(args.relatorio).resolve() if args.relatorio else None
    if args.somente_classificar:
        analisar_local(registros, relatorio)
        return

    config = carregar_config(Path(args.config).resolve())
    glpi = Glpi(config)
    resultados = []
    alterados = 0
    simulados = 0
    ja_categorizados = 0
    ausentes = 0
    erros = 0
    intervalo = float(config.get("interval_seconds", 0.25))

    try:
        glpi.iniciar()
        categorias_glpi = mapa_categorias_glpi(glpi.listar_categorias(), config)
        categorias_necessarias = {
            normalizar(item["categoria"]): item["categoria"]
            for item in registros
            if item["categoria"] and (args.incluir_a_classificar or normalizar(item["categoria"]) != normalizar(CATEGORIA_A_CLASSIFICAR))
        }
        faltantes = sorted(
            nome_original for chave, nome_original in categorias_necessarias.items()
            if chave not in categorias_glpi
        )
        if faltantes:
            print("ERRO: estas categorias da planilha não foram encontradas no GLPI:")
            for nome in faltantes:
                print(f"  - {nome}")
            print("Nenhum chamado foi alterado. Crie as categorias ou informe seus IDs em category_map.")
            raise SystemExit(2)

        modo = "EXECUÇÃO" if args.executar else "SIMULAÇÃO"
        limite_texto = "todos" if args.limit == 0 else str(args.limit)
        print(f"{modo}: {len(registros)} linhas disponíveis; limite de {limite_texto} alterações.")

        for item in registros:
            if args.limit > 0 and (alterados + simulados) >= args.limit:
                break
            categoria = item["categoria"]
            if normalizar(categoria) == normalizar(CATEGORIA_A_CLASSIFICAR) and not args.incluir_a_classificar:
                resultados.append({
                    "ID": item["id"], "Título": item["titulo"], "Categoria": categoria,
                    "Origem": item["origem_categoria"], "Resultado": "IGNORADO",
                    "Detalhe": "Categoria A classificar exige --incluir-a-classificar",
                })
                continue

            ticket = glpi.obter_ticket(item["id"])
            if ticket is None:
                ausentes += 1
                print(f"AUSENTE: chamado GLPI #{item['id']} não encontrado.")
                resultados.append({
                    "ID": item["id"], "Título": item["titulo"], "Categoria": categoria,
                    "Origem": item["origem_categoria"], "Resultado": "AUSENTE", "Detalhe": "Chamado não encontrado no GLPI",
                })
                continue
            atual = categoria_atual_id(ticket)
            if atual:
                ja_categorizados += 1
                print(f"JÁ POSSUI: GLPI #{item['id']} já possui categoria ID {atual}; nenhuma alteração.")
                resultados.append({
                    "ID": item["id"], "Título": item["titulo"], "Categoria": categoria,
                    "Origem": item["origem_categoria"], "Resultado": "JÁ POSSUI", "Detalhe": f"Categoria atual ID {atual}",
                })
                continue

            categoria_id = categorias_glpi[normalizar(categoria)]
            if args.executar:
                try:
                    glpi.atribuir_categoria(item["id"], categoria_id)
                    confirmacao = glpi.obter_ticket(item["id"])
                    if not confirmacao or categoria_atual_id(confirmacao) != categoria_id:
                        raise RuntimeError("A API respondeu, mas a categoria não foi confirmada na releitura")
                    alterados += 1
                    print(f"OK: GLPI #{item['id']} -> {categoria} (ID {categoria_id})")
                    resultado, detalhe = "ATUALIZADO", f"Categoria ID {categoria_id} confirmada"
                except Exception as erro:
                    erros += 1
                    print(f"ERRO: GLPI #{item['id']} - {erro}")
                    resultado, detalhe = "ERRO", str(erro)
            else:
                simulados += 1
                print(f"SIMULAR: GLPI #{item['id']} -> {categoria} (ID {categoria_id})")
                resultado, detalhe = "SIMULAÇÃO", f"Categoria ID {categoria_id}; nenhuma alteração realizada"

            resultados.append({
                "ID": item["id"], "Título": item["titulo"], "Categoria": categoria,
                "Origem": item["origem_categoria"], "Resultado": resultado, "Detalhe": detalhe,
            })
            if intervalo > 0:
                time.sleep(intervalo)
    finally:
        glpi.finalizar()
        if relatorio:
            salvar_relatorio(relatorio, resultados)

    print(
        f"RESUMO: {alterados} atualizados; {simulados} simulados; "
        f"{ja_categorizados} já categorizados; {ausentes} ausentes; {erros} erros."
    )
    if relatorio:
        print(f"Relatório: {relatorio}")
    if erros:
        raise SystemExit(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Operação cancelada pelo usuário.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as erro:
        print(f"ERRO FATAL: {erro}", file=sys.stderr)
        raise SystemExit(1)
