#!/usr/bin/env python3
"""Verifica planilha, configuração e controle ANTES de qualquer chamado ser
criado ou alterado no GLPI.

Este módulo nunca cria, atualiza ou apaga nada no GLPI. Com --api, ele
apenas faz consultas de leitura (initSession, GET) para confirmar que a
API, os tokens e os IDs configurados realmente existem no servidor.

Se qualquer verificação crítica falhar, o script termina com código de
saída diferente de zero. O importador_geral.py trata isso como erro e
interrompe o fluxo antes de chegar perto do GLPI.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", message="Cell .* is marked as a date.*", category=UserWarning)

COLUNAS_OBRIGATORIAS = [
    "Carimbo de data/hora",
    "Nome do Solicitante",
    "Departamento",
    "Prioridade",
    "Assunto",
    "Motivo da Solicitação",
    "Descrição do Chamado",
    "Endereço de e-mail",
    "Responsável",
    "Status",
]
COLUNA_ANEXO_1 = 12  # L
COLUNA_ANEXO_2 = 14  # N
TRECHO_CABECALHO_ANEXO = "foto do erro"
PRIORIDADES_CONHECIDAS = {"baixa", "normal", "urgente"}
STATUS_CONHECIDOS = {"finalizado", "suspenso", "em andamento", "aguardando validação"}
REGEX_REFERENCIA = re.compile(r"^EXCEL-(\d+)$", re.IGNORECASE)

# A coluna de data de abertura já apareceu com nomes diferentes em
# exportações diferentes do Microsoft Forms; qualquer um destes é aceito
# e tratado internamente como "Carimbo de data/hora".
ALIASES_COLUNA_ABERTURA = {"carimbo de data/hora", "data de abertura"}


def normalizar_colunas(df):
    renomear = {}
    for coluna in df.columns:
        if str(coluna).strip().casefold() in ALIASES_COLUNA_ABERTURA:
            renomear[coluna] = "Carimbo de data/hora"
    return df.rename(columns=renomear) if renomear else df


def texto(valor):
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    return str(valor).strip()


def chave(valor):
    return texto(valor).casefold()


def linha_valida(linha):
    tem_data = bool(texto(linha.get("Carimbo de data/hora")))
    tem_solicitante = bool(texto(linha.get("Nome do Solicitante")))
    tem_descricao = bool(
        texto(linha.get("Descrição do Chamado")) or texto(linha.get("Motivo da Solicitação"))
    )
    return tem_data and tem_solicitante and tem_descricao


class Relatorio:
    def __init__(self):
        self.linhas = []
        self.erros = 0
        self.avisos = 0

    def secao(self, titulo):
        self.linhas.append("")
        self.linhas.append(f"== {titulo} ==")

    def ok(self, msg):
        self.linhas.append(f"[OK]    {msg}")

    def aviso(self, msg):
        self.linhas.append(f"[AVISO] {msg}")
        self.avisos += 1

    def erro(self, msg):
        self.linhas.append(f"[ERRO]  {msg}")
        self.erros += 1

    def texto_final(self):
        cabecalho = [
            "VERIFICAÇÃO PRÉ-IMPORTAÇÃO GLPI",
            f"Erros: {self.erros}  |  Avisos: {self.avisos}",
        ]
        return "\n".join(cabecalho + self.linhas) + "\n"


def limitar(itens, maximo=10):
    itens = list(itens)
    if len(itens) <= maximo:
        return ", ".join(str(i) for i in itens)
    resto = len(itens) - maximo
    return ", ".join(str(i) for i in itens[:maximo]) + f" (+{resto} outros)"


# --------------------------------------------------------------------------
# 1. Arquivos básicos
# --------------------------------------------------------------------------

def verificar_arquivos(config_path, estado_path, planilha_path, rel):
    rel.secao("Arquivos")
    for rotulo, caminho, obrigatorio in (
        ("config.json", config_path, True),
        ("planilha", planilha_path, True),
        ("estado (controle EXCEL-N)", estado_path, False),
    ):
        if not caminho.exists():
            (rel.erro if obrigatorio else rel.aviso)(f"{rotulo} não encontrado: {caminho}")
        elif caminho.stat().st_size == 0:
            rel.erro(f"{rotulo} está vazio: {caminho}")
        else:
            rel.ok(f"{rotulo} encontrado ({caminho.stat().st_size} bytes).")


# --------------------------------------------------------------------------
# 2. config.json
# --------------------------------------------------------------------------

def verificar_config(config_path, rel):
    if not config_path.exists():
        return None
    try:
        config = json.loads(config_path.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError as erro:
        rel.erro(f"config.json não é um JSON válido: {erro}")
        return None

    rel.secao("Configuração (config.json)")

    for campo in ("api_url", "app_token", "user_token"):
        valor = texto(config.get(campo))
        if not valor:
            rel.erro(f"Campo obrigatório vazio no config.json: {campo!r}.")

    api_url = texto(config.get("api_url"))
    if api_url:
        if not re.match(r"^https?://", api_url):
            rel.erro(
                f"api_url inválida: {api_url!r}. Deve começar exatamente com "
                "'http://' ou 'https://', sem texto extra antes do endereço "
                "(ex.: não use 'URL: https://...')."
            )
        elif api_url != api_url.strip():
            rel.aviso("api_url tem espaços no início/fim; será usada mesmo assim.")
        else:
            rel.ok(f"api_url válida: {api_url}")

    if texto(config.get("app_token")) and texto(config.get("user_token")):
        rel.ok("app_token e user_token estão preenchidos.")

    intervalo = config.get("interval_seconds", 0.3)
    try:
        intervalo = float(intervalo)
        if intervalo < 0.05:
            rel.aviso(f"interval_seconds muito baixo ({intervalo}s); risco de bloqueio por excesso de requisições.")
        else:
            rel.ok(f"interval_seconds = {intervalo}s.")
    except (TypeError, ValueError):
        rel.erro(f"interval_seconds inválido: {config.get('interval_seconds')!r}.")

    tamanho_pagina = config.get("duplicate_scan_page_size", 500)
    try:
        tamanho_pagina = int(tamanho_pagina)
        if not (20 <= tamanho_pagina <= 500):
            rel.aviso(f"duplicate_scan_page_size fora da faixa recomendada (20-500): {tamanho_pagina}.")
    except (TypeError, ValueError):
        rel.erro(f"duplicate_scan_page_size inválido: {config.get('duplicate_scan_page_size')!r}.")

    for chave_mapa in ("technician_map", "category_map", "department_group_map", "requester_user_map"):
        valor = config.get(chave_mapa, {})
        if not isinstance(valor, dict):
            rel.erro(f"{chave_mapa} deveria ser um objeto (dicionário) no config.json.")
        else:
            rel.ok(f"{chave_mapa}: {len(valor)} entrada(s).")

    for nome_mapa in ("technician_map", "requester_user_map"):
        mapa = config.get(nome_mapa, {})
        if not isinstance(mapa, dict):
            continue
        ids_nao_numericos = [k for k, v in mapa.items() if not str(v).strip().lstrip("-").isdigit()]
        if ids_nao_numericos:
            rel.erro(f"{nome_mapa} tem valores que não são IDs numéricos: {limitar(ids_nao_numericos)}.")

    return config


# --------------------------------------------------------------------------
# 3. Estado (controle EXCEL-N -> ticket GLPI)
# --------------------------------------------------------------------------

def verificar_estado(estado_path, rel):
    if not estado_path.exists():
        rel.secao("Controle (importacao_glpi_estado.json)")
        rel.aviso("Arquivo de controle ainda não existe; será criado na primeira importação.")
        return {"importados": {}}
    try:
        estado = json.loads(estado_path.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError as erro:
        rel.erro(f"Arquivo de controle não é um JSON válido: {erro}")
        return None

    rel.secao("Controle (importacao_glpi_estado.json)")
    if not isinstance(estado, dict) or not isinstance(estado.get("importados", {}), dict):
        rel.erro("Estrutura do controle é inválida (esperado um objeto com a chave 'importados').")
        return None

    importados = estado["importados"]
    rel.ok(f"{len(importados)} referência(s) já registradas no controle.")

    invalidas = [ref for ref in importados if not REGEX_REFERENCIA.match(ref)]
    if invalidas:
        rel.erro(f"Referências fora do padrão EXCEL-N no controle: {limitar(invalidas)}.")

    por_ticket = defaultdict(list)
    sem_ticket_id = []
    for ref, dados in importados.items():
        ticket_id = dados.get("ticket_id") if isinstance(dados, dict) else None
        if ticket_id is None:
            sem_ticket_id.append(ref)
        else:
            por_ticket[int(ticket_id)].append(ref)

    if sem_ticket_id:
        rel.erro(f"Referências sem ticket_id no controle: {limitar(sem_ticket_id)}.")

    duplicados = {tid: refs for tid, refs in por_ticket.items() if len(refs) > 1}
    if duplicados:
        exemplos = limitar([f"GLPI #{tid} <- {', '.join(refs)}" for tid, refs in duplicados.items()], 5)
        rel.erro(
            f"{len(duplicados)} chamado(s) do GLPI estão vinculados a mais de uma referência EXCEL-N "
            f"(risco de sobrescrever o chamado errado): {exemplos}."
        )
    else:
        rel.ok("Nenhum ticket_id duplicado entre as referências do controle.")

    return estado


# --------------------------------------------------------------------------
# 4. Planilha
# --------------------------------------------------------------------------

def carregar_links_anexos(caminho_excel, rel):
    planilha = load_workbook(caminho_excel, read_only=False, data_only=False).active
    cabecalho_1 = texto(planilha.cell(1, COLUNA_ANEXO_1).value).casefold()
    cabecalho_2 = texto(planilha.cell(1, COLUNA_ANEXO_2).value).casefold()
    if TRECHO_CABECALHO_ANEXO not in cabecalho_1 or TRECHO_CABECALHO_ANEXO not in cabecalho_2:
        rel.erro(
            "As colunas de anexos parecem estar deslocadas: era esperado encontrar "
            f"'{TRECHO_CABECALHO_ANEXO}' nas colunas L e N. Encontrado: "
            f"L={planilha.cell(1, COLUNA_ANEXO_1).value!r}, N={planilha.cell(1, COLUNA_ANEXO_2).value!r}. "
            "Não altere a posição dessas colunas na planilha."
        )
    else:
        rel.ok("Colunas de anexos (L e N) estão na posição esperada.")

    anexos = {}
    total_links = 0
    for linha in range(2, planilha.max_row + 1):
        itens = []
        for coluna in (COLUNA_ANEXO_1, COLUNA_ANEXO_2):
            celula = planilha.cell(linha, coluna)
            if celula.hyperlink and celula.value:
                itens.append((texto(celula.value), celula.hyperlink.target))
        if itens:
            anexos[linha] = itens
            total_links += len(itens)
    rel.ok(f"{len(anexos)} linha(s) com anexo(s); {total_links} link(s) de anexo no total.")
    return anexos


def verificar_planilha(planilha_path, rel):
    rel.secao("Planilha")
    try:
        df = pd.read_excel(planilha_path)
        df = normalizar_colunas(df)
    except Exception as erro:
        rel.erro(f"Não foi possível abrir a planilha: {erro}")
        return None, {}, None

    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]
    if faltando:
        rel.erro(f"Colunas obrigatórias ausentes na planilha: {', '.join(faltando)}.")
        return None, {}, None
    rel.ok(f"Planilha aberta com {len(df)} linha(s) e todas as colunas obrigatórias presentes.")

    # O importador (importar_chamados.py) descarta linhas com Nome do
    # Solicitante = "Ignorar" antes de processar; a verificação precisa
    # fazer exatamente o mesmo, senão relata problemas em linhas de
    # template que nunca seriam importadas.
    total_antes = len(df)
    df = df[df["Nome do Solicitante"].fillna("").astype(str).str.strip().str.casefold() != "ignorar"]
    ignoradas = total_antes - len(df)
    if ignoradas:
        rel.ok(
            f"{ignoradas} linha(s) com Nome do Solicitante = 'Ignorar' foram excluídas desta "
            "verificação, do mesmo jeito que o importador as ignora."
        )

    validas = 0
    invalidas = []
    for indice, linha in df.iterrows():
        if linha_valida(linha):
            validas += 1
        elif any(texto(linha.get(c)) for c in df.columns):
            invalidas.append(f"EXCEL-{indice + 2}")
    rel.ok(f"{validas} linha(s) válida(s) para importação.")
    if invalidas:
        rel.aviso(
            f"{len(invalidas)} linha(s) com algum dado preenchido mas incompletas "
            f"(sem data, solicitante ou descrição) - serão ignoradas: {limitar(invalidas)}."
        )

    datas_invalidas = []
    for indice, linha in df.iterrows():
        if not linha_valida(linha):
            continue
        bruto = linha.get("Carimbo de data/hora")
        try:
            dt = pd.to_datetime(bruto, errors="raise", dayfirst=True)
            if dt.year < 2023 or dt > pd.Timestamp.now() + pd.Timedelta(1, unit="D"):
                datas_invalidas.append(f"EXCEL-{indice + 2}")
        except Exception:
            datas_invalidas.append(f"EXCEL-{indice + 2}")
    if datas_invalidas:
        rel.aviso(
            f"{len(datas_invalidas)} linha(s) com 'Carimbo de data/hora' vazio, ilegível ou fora do "
            f"intervalo esperado (o chamado será criado sem data de abertura definida): {limitar(datas_invalidas)}."
        )
    else:
        rel.ok("Todas as datas de abertura são legíveis e plausíveis.")

    if "Data da Finalização" in df.columns:
        datas_fim_invalidas = []
        for indice, linha in df.iterrows():
            if not linha_valida(linha):
                continue
            status = chave(linha.get("Status"))
            bruto = linha.get("Data da Finalização")
            if status not in {"finalizado", "aguardando validação"} or not texto(bruto):
                continue
            try:
                dt = pd.to_datetime(bruto, errors="raise", dayfirst=True)
                if dt.year < 2023 or dt > pd.Timestamp.now() + pd.Timedelta(1, unit="D"):
                    datas_fim_invalidas.append(f"EXCEL-{indice + 2}")
            except Exception:
                datas_fim_invalidas.append(f"EXCEL-{indice + 2}")
        if datas_fim_invalidas:
            rel.aviso(
                f"{len(datas_fim_invalidas)} chamado(s) finalizado(s)/aguardando validação com "
                f"'Data da Finalização' ilegível: {limitar(datas_fim_invalidas)}."
            )

    prioridades_desconhecidas = sorted({
        texto(v) for v in df["Prioridade"] if texto(v) and chave(v) not in PRIORIDADES_CONHECIDAS
    })
    if prioridades_desconhecidas:
        rel.aviso(
            f"Valores de Prioridade não mapeados (serão tratados como 'normal'): "
            f"{limitar(prioridades_desconhecidas)}."
        )

    status_desconhecidos = sorted({
        texto(v) for v in df["Status"] if texto(v) and chave(v) not in STATUS_CONHECIDOS
    })
    if status_desconhecidos:
        rel.aviso(
            f"Valores de Status não mapeados (serão tratados como 'Novo'): "
            f"{limitar(status_desconhecidos)}."
        )

    links_anexos = carregar_links_anexos(planilha_path, rel)
    return df, links_anexos, total_antes


# --------------------------------------------------------------------------
# 5. Comparação com a planilha anterior (linhas antigas não podem mudar)
# --------------------------------------------------------------------------

def verificar_planilha_anterior(planilha_path, df, total_linhas_atual, estado, rel):
    rel.secao("Comparação com a planilha anterior")
    candidata = planilha_path.parent / "planilha_anterior.xlsx"
    if not candidata.exists() or candidata == planilha_path:
        rel.aviso("planilha_anterior.xlsx não encontrada ao lado da planilha atual; comparação pulada.")
        return
    try:
        df_anterior = pd.read_excel(candidata)
        df_anterior = normalizar_colunas(df_anterior)
    except Exception as erro:
        rel.erro(f"Não foi possível abrir planilha_anterior.xlsx: {erro}")
        return

    if len(df_anterior) > total_linhas_atual:
        rel.erro(
            f"planilha_anterior.xlsx tem mais linhas ({len(df_anterior)}) do que a planilha atual "
            f"({total_linhas_atual}). Isso indica que linhas antigas podem ter sido removidas - a "
            "referência EXCEL-N depende da posição original da linha."
        )
        return

    # df pode ter linhas 'Ignorar' removidas: usar .loc (por rótulo/posição
    # original) em vez de .iloc (posicional), senão a comparação desalinha.
    importados = (estado or {}).get("importados", {})
    divergencias = []
    for ref in importados:
        m = REGEX_REFERENCIA.match(ref)
        if not m:
            continue
        indice = int(m.group(1)) - 2
        if indice < 0 or indice >= len(df_anterior) or indice not in df.index:
            continue
        antigo = df_anterior.iloc[indice]
        atual = df.loc[indice]
        if chave(antigo.get("Nome do Solicitante")) != chave(atual.get("Nome do Solicitante")) or texto(
            antigo.get("Carimbo de data/hora")
        ) != texto(atual.get("Carimbo de data/hora")):
            divergencias.append(ref)

    if divergencias:
        rel.aviso(
            f"{len(divergencias)} referência(s) já importada(s) mudaram de solicitante/data entre a "
            "planilha anterior e a atual. Isso é esperado quando uma linha antes vazia foi preenchida "
            "depois (use a opção 'reparar' do módulo de importação para essas referências); se não for "
            f"o caso, confira se as linhas foram reordenadas ou excluídas: {limitar(divergencias)}."
        )
    else:
        rel.ok("Linhas já importadas continuam na mesma posição e com os mesmos dados-chave.")


# --------------------------------------------------------------------------
# 6. Referências (EXCEL-N) vs. planilha atual
# --------------------------------------------------------------------------

def verificar_referencias(total_linhas, estado, rel):
    rel.secao("Referências EXCEL-N")
    if total_linhas is None or estado is None:
        rel.aviso("Verificação de referências pulada por falha em uma etapa anterior.")
        return

    importados = estado.get("importados", {})
    fora_da_planilha = []
    for ref in importados:
        m = REGEX_REFERENCIA.match(ref)
        if not m:
            continue
        numero = int(m.group(1))
        indice = numero - 2
        if indice < 0 or indice >= total_linhas:
            fora_da_planilha.append(ref)

    if fora_da_planilha:
        rel.erro(
            f"{len(fora_da_planilha)} referência(s) do controle apontam para linhas que não existem "
            f"mais na planilha atual (ela pode ter encolhido): {limitar(fora_da_planilha)}."
        )
    else:
        rel.ok("Todas as referências do controle correspondem a linhas existentes na planilha atual.")


# --------------------------------------------------------------------------
# 7. Cobertura dos mapas (técnico / requerente)
# --------------------------------------------------------------------------

def linhas_pendentes(df, estado):
    """Só as linhas válidas que ainda NÃO foram importadas. As checagens de
    cobertura de mapa e de IDs no GLPI (--api) usam isso, não a planilha
    inteira: um valor usado só em referências já importadas não deveria
    bloquear a importação das linhas novas."""
    importados = (estado or {}).get("importados", {})
    indices = [
        indice for indice, linha in df.iterrows()
        if linha_valida(linha) and f"EXCEL-{indice + 2}" not in importados
    ]
    return df.loc[indices]


def verificar_cobertura_mapas(df, estado, config, rel):
    rel.secao("Cobertura dos mapas de técnico, requerente, categoria e departamento")
    if df is None or config is None:
        rel.aviso("Verificação de cobertura pulada por falha em uma etapa anterior.")
        return []

    df = linhas_pendentes(df, estado)
    if df.empty:
        rel.ok("Nenhuma linha pendente de importação; verificação de cobertura não se aplica.")
        return []

    tecnicos_map = {chave(k) for k in config.get("technician_map", {})}
    requerentes_map = {chave(k) for k in config.get("requester_user_map", {})}
    categorias_map = {chave(k) for k in config.get("category_map", {})}
    departamentos_map = {chave(k) for k in config.get("department_group_map", {})}
    auto_cria_categorias = bool(config.get("auto_create_categories", False))

    responsaveis = Counter(texto(v) for v in df["Responsável"] if texto(v))
    faltando_tecnico = sorted(v for v in responsaveis if chave(v) not in tecnicos_map)
    if faltando_tecnico:
        rel.aviso(
            f"{len(faltando_tecnico)} valor(es) de 'Responsável' sem correspondência em technician_map "
            f"(o chamado será criado sem técnico atribuído): {limitar(faltando_tecnico)}."
        )
    else:
        rel.ok("Todos os responsáveis das linhas pendentes de importação estão no technician_map.")

    nomes_completos = (
        df["Nome do Solicitante"].fillna("") + " " + df.get("Sobrenome do Solicitante", "").fillna("")
    ).str.strip()
    solicitantes = Counter(v for v in nomes_completos if v)
    faltando_requerente = sorted(v for v in solicitantes if chave(v) not in requerentes_map)
    if faltando_requerente:
        rel.aviso(
            f"{len(faltando_requerente)} solicitante(s) sem correspondência em requester_user_map "
            "(use a opção 'Validar usuários requerentes' para resolver antes de rodar em massa): "
            f"{limitar(faltando_requerente, 8)}."
        )
    else:
        rel.ok("Todos os solicitantes das linhas pendentes de importação estão no requester_user_map.")

    faltando_categoria = []
    if "Assunto" in df.columns:
        assuntos = Counter(texto(v) for v in df["Assunto"] if texto(v))
        faltando_categoria = sorted(v for v in assuntos if chave(v) not in categorias_map)
        if faltando_categoria and not auto_cria_categorias:
            rel.aviso(
                f"{len(faltando_categoria)} valor(es) de 'Assunto' sem entrada em category_map "
                "(auto_create_categories está desligado): o importador ainda tenta achar uma categoria "
                "com esse MESMO NOME já existente no GLPI antes de deixar sem categoria - isso só é "
                "confirmado com --api (verificação real no servidor). Sem --api, trate como pendente: "
                f"{limitar(faltando_categoria)}."
            )
        elif faltando_categoria:
            rel.aviso(
                f"{len(faltando_categoria)} valor(es) de 'Assunto' sem entrada em category_map, "
                f"mas serão criados automaticamente se não existirem no GLPI (auto_create_categories=true): "
                f"{limitar(faltando_categoria)}."
            )
        else:
            rel.ok("Todos os assuntos das linhas pendentes de importação estão no category_map.")

    if "Departamento" in df.columns:
        departamentos = Counter(texto(v) for v in df["Departamento"] if texto(v))
        faltando_departamento = sorted(v for v in departamentos if chave(v) not in departamentos_map)
        if faltando_departamento:
            rel.aviso(
                f"{len(faltando_departamento)} valor(es) de 'Departamento' sem correspondência em "
                "department_group_map (o chamado é criado, mas o passo 'Completar chamados existentes' "
                f"não vai conseguir atribuir o grupo requerente): {limitar(faltando_departamento)}."
            )
        else:
            rel.ok("Todos os departamentos das linhas pendentes de importação estão no department_group_map.")

    return faltando_categoria


# --------------------------------------------------------------------------
# 7c. Solução ausente em chamados finalizados
# --------------------------------------------------------------------------

def verificar_solucao_finalizados(df, rel):
    rel.secao("Solução dos chamados finalizados")
    if df is None or "Desdobramento" not in df.columns:
        rel.aviso("Coluna 'Desdobramento' não encontrada; verificação de solução pulada.")
        return

    sem_solucao = []
    for indice, linha in df.iterrows():
        if not linha_valida(linha):
            continue
        status = chave(linha.get("Status"))
        if status not in {"finalizado", "aguardando validação"}:
            continue
        if not texto(linha.get("Desdobramento")):
            sem_solucao.append(f"EXCEL-{indice + 2}")

    if sem_solucao:
        rel.aviso(
            f"{len(sem_solucao)} chamado(s) finalizado(s)/aguardando validação SEM texto em "
            "'Desdobramento'. O importador não deixa o campo solução vazio: preenche com um texto "
            "genérico ('Chamado concluído no controle antigo; a solução não foi registrada no Excel.') "
            f"em vez da solução real. Revise se vale preencher a coluna antes de importar: "
            f"{limitar(sem_solucao, 10)}."
        )
    else:
        rel.ok("Todos os chamados finalizados/aguardando validação têm texto em 'Desdobramento'.")


# --------------------------------------------------------------------------
# 7b. Chamados repetidos (mesmo solicitante + mesma descrição)
# --------------------------------------------------------------------------

def verificar_chamados_repetidos(df, estado, rel):
    rel.secao("Chamados possivelmente repetidos")
    if df is None:
        rel.aviso("Verificação de repetidos pulada por falha em uma etapa anterior.")
        return

    importados = (estado or {}).get("importados", {})
    grupos = defaultdict(list)
    for indice, linha in df.iterrows():
        numero = f"EXCEL-{indice + 2}"
        if not linha_valida(linha) or numero in importados:
            continue
        nome = " ".join(filter(None, [
            texto(linha.get("Nome do Solicitante")),
            texto(linha.get("Sobrenome do Solicitante")),
        ]))
        descricao = texto(linha.get("Descrição do Chamado")) or texto(linha.get("Motivo da Solicitação"))
        descricao_normalizada = re.sub(r"\s+", " ", descricao).strip().casefold()
        if not descricao_normalizada:
            continue
        grupos[(chave(nome), descricao_normalizada)].append((numero, texto(linha.get("Carimbo de data/hora"))))

    repetidos = {k: v for k, v in grupos.items() if len(v) > 1}
    if not repetidos:
        rel.ok("Nenhum chamado ainda não importado com o mesmo solicitante e a mesma descrição.")
        return

    total_linhas = sum(len(v) for v in repetidos.values())
    exemplos = []
    for (_, _), linhas in repetidos.items():
        refs = ", ".join(f"{ref} ({data[:10]})" if data else ref for ref, data in linhas)
        exemplos.append(refs)
    rel.aviso(
        f"{len(repetidos)} grupo(s) de chamados repetidos ({total_linhas} linha(s) no total): mesmo "
        "solicitante e mesma descrição em mais de uma referência EXCEL-N ainda não importada. O "
        "importador (importar_chamados.py) já trata isso por padrão: só cria o chamado da primeira "
        "referência do grupo e vincula as demais ao mesmo chamado no GLPI, sem duplicar. Revise a "
        f"lista antes de importar em massa (use --permitir-duplicados se algum destes for legítimo): "
        f"{limitar(exemplos, 8)}."
    )


# --------------------------------------------------------------------------
# 8. Anexos locais
# --------------------------------------------------------------------------

def verificar_anexos_locais(links_anexos, config, raiz, rel):
    rel.secao("Anexos")
    pasta = raiz / texto(config.get("local_attachment_dir", "anexos_forms")) if config else raiz / "anexos_forms"
    if not pasta.exists():
        rel.aviso(f"Pasta de anexos locais não existe ainda: {pasta}. Anexos serão baixados do link original.")
        return

    nomes_locais = {p.name.casefold() for p in pasta.rglob("*") if p.is_file()}
    sem_correspondencia_local = 0
    total = 0
    for itens in links_anexos.values():
        for nome, _url in itens:
            total += 1
            if Path(nome).name.casefold() not in nomes_locais:
                sem_correspondencia_local += 1
    if total:
        rel.ok(
            f"{total - sem_correspondencia_local} de {total} anexo(s) já têm cópia local em "
            f"{pasta.name}; os demais serão baixados do link (ex.: Google Drive) durante a importação."
        )


# --------------------------------------------------------------------------
# 9. API (opcional, somente leitura)
# --------------------------------------------------------------------------

def _requisicao(api_url, headers, metodo, endpoint, timeout=30):
    req = urllib.request.Request(f"{api_url}/{endpoint.lstrip('/')}", headers=headers, method=metodo)
    with urllib.request.urlopen(req, timeout=timeout) as resposta:
        conteudo = resposta.read().decode("utf-8")
    return json.loads(conteudo) if conteudo else None


def verificar_api(config, assuntos_sem_map, chaves_relevantes, rel):
    rel.secao("API do GLPI (somente leitura)")
    if not config:
        rel.aviso("Verificação de API pulada: configuração inválida.")
        return
    api_url = texto(config.get("api_url")).rstrip("/")
    if not re.match(r"^https?://", api_url):
        rel.erro("Verificação de API pulada: api_url inválida (ver seção Configuração).")
        return

    headers = {
        "Content-Type": "application/json",
        "App-Token": texto(config.get("app_token")),
        "Authorization": f"user_token {texto(config.get('user_token'))}",
    }
    session_token = None
    try:
        resposta = _requisicao(api_url, headers, "GET", "initSession", timeout=20)
        session_token = resposta["session_token"]
        headers["Session-Token"] = session_token
        rel.ok("Login na API (initSession) funcionou com os tokens do config.json.")

        perfil_id = config.get("profile_id", 0)
        if perfil_id:
            try:
                req = urllib.request.Request(
                    f"{api_url}/changeActiveProfile",
                    data=json.dumps({"profiles_id": int(perfil_id)}).encode("utf-8"),
                    headers=headers,
                    method="POST",
                )
                urllib.request.urlopen(req, timeout=20)
                rel.ok(f"Perfil ativo alterado para profile_id={perfil_id}.")
            except urllib.error.HTTPError as erro:
                rel.erro(f"profile_id={perfil_id} não foi aceito pelo GLPI (HTTP {erro.code}).")

        try:
            _requisicao(api_url, headers, "GET", "Ticket?range=0-0", timeout=20)
            rel.ok("Permissão de leitura em Ticket confirmada.")
        except urllib.error.HTTPError as erro:
            rel.erro(f"Sem permissão de leitura em Ticket: HTTP {erro.code}.")

        for nome_mapa, endpoint in (("technician_map", "User"), ("requester_user_map", "User")):
            chaves_pendentes = chaves_relevantes.get(nome_mapa)
            itens_mapa = config.get(nome_mapa, {}).items()
            if chaves_pendentes is not None:
                # Só interessa, para bloquear a importação, o ID de quem
                # ainda vai ser usado em alguma linha pendente. Uma
                # referência antiga, usada só por chamados já importados,
                # não deveria travar a importação das linhas novas.
                itens_mapa = [(k, v) for k, v in itens_mapa if chave(k) in chaves_pendentes]
            ids = sorted({int(v) for _, v in itens_mapa if str(v).strip().lstrip('-').isdigit()})
            invalidos = []
            for item_id in ids:
                try:
                    retorno = _requisicao(api_url, headers, "GET", f"{endpoint}/{item_id}", timeout=15)
                    if not isinstance(retorno, dict) or int(retorno.get("id") or 0) != item_id:
                        invalidos.append(item_id)
                except urllib.error.HTTPError:
                    invalidos.append(item_id)
            if invalidos:
                rel.erro(
                    f"{nome_mapa}: ID(s) de usuário inexistente(s) no GLPI e usado(s) em alguma linha "
                    f"pendente de importação: {limitar(invalidos)}."
                )
            elif ids:
                rel.ok(f"{nome_mapa}: os {len(ids)} ID(s) usados nas linhas pendentes existem no GLPI.")
            else:
                rel.ok(f"{nome_mapa}: nenhuma linha pendente depende de um ID mapeado aqui.")

        ids_categorias = sorted({
            int(v) for v in config.get("category_map", {}).values() if str(v).strip().lstrip('-').isdigit()
        })
        invalidas = []
        for item_id in ids_categorias:
            try:
                retorno = _requisicao(api_url, headers, "GET", f"ITILCategory/{item_id}", timeout=15)
                if not isinstance(retorno, dict) or int(retorno.get("id") or 0) != item_id:
                    invalidas.append(item_id)
            except urllib.error.HTTPError:
                invalidas.append(item_id)
        if invalidas:
            rel.erro(f"category_map: ID(s) de categoria inexistente(s) no GLPI: {limitar(invalidas)}.")
        elif ids_categorias:
            rel.ok(f"category_map: todos os {len(ids_categorias)} ID(s) de categoria existem no GLPI.")

        if assuntos_sem_map:
            try:
                categorias_glpi = _requisicao(api_url, headers, "GET", "ITILCategory?range=0-9999", timeout=30)
                nomes_glpi = {
                    texto(c.get("completename") or c.get("name")).casefold()
                    for c in (categorias_glpi if isinstance(categorias_glpi, list) else [])
                }
                sem_categoria_de_verdade = sorted(a for a in assuntos_sem_map if a.casefold() not in nomes_glpi)
                if sem_categoria_de_verdade and not config.get("auto_create_categories", False):
                    rel.erro(
                        f"{len(sem_categoria_de_verdade)} valor(es) de 'Assunto' NÃO têm categoria com "
                        "esse nome no GLPI nem entrada no category_map, e auto_create_categories está "
                        f"desligado: esses chamados serão criados SEM categoria: {limitar(sem_categoria_de_verdade)}."
                    )
                elif not sem_categoria_de_verdade:
                    rel.ok(
                        "Todos os 'Assunto' sem entrada no category_map já existem como categoria no "
                        "GLPI com o mesmo nome (serão casados automaticamente)."
                    )
            except urllib.error.HTTPError as erro:
                rel.erro(f"Não foi possível confirmar as categorias existentes no GLPI: HTTP {erro.code}.")

    except urllib.error.HTTPError as erro:
        corpo = erro.read().decode("utf-8", errors="replace")[:400]
        rel.erro(f"Falha ao autenticar na API do GLPI: HTTP {erro.code} - {corpo}")
    except urllib.error.URLError as erro:
        rel.erro(f"Não foi possível conectar à API do GLPI ({api_url}): {erro.reason}")
    except Exception as erro:  # noqa: BLE001
        rel.erro(f"Erro inesperado ao testar a API: {erro}")
    finally:
        if session_token:
            try:
                _requisicao(api_url, headers, "GET", "killSession", timeout=15)
            except Exception:
                pass


# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Verificação completa antes de importar/alterar chamados no GLPI.")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--estado", default="importacao_glpi_estado.json")
    parser.add_argument("--planilha", default="planilha_atualizada.xlsx")
    parser.add_argument("--relatorio", default="verificacao_pre_importacao.txt")
    parser.add_argument("--api", action="store_true", help="Também testa a API do GLPI (somente leitura)")
    args = parser.parse_args()

    config_path = Path(args.config).resolve()
    estado_path = Path(args.estado).resolve()
    planilha_path = Path(args.planilha).resolve()
    relatorio_path = Path(args.relatorio).resolve()

    rel = Relatorio()
    verificar_arquivos(config_path, estado_path, planilha_path, rel)
    config = verificar_config(config_path, rel)
    estado = verificar_estado(estado_path, rel)
    df, links_anexos, total_linhas_planilha = (
        verificar_planilha(planilha_path, rel) if planilha_path.exists() else (None, {}, None)
    )
    assuntos_sem_map = []
    chaves_relevantes = {"technician_map": set(), "requester_user_map": set()}
    if df is not None:
        verificar_planilha_anterior(planilha_path, df, total_linhas_planilha, estado, rel)
        verificar_referencias(total_linhas_planilha, estado, rel)
        verificar_chamados_repetidos(df, estado, rel)
        assuntos_sem_map = verificar_cobertura_mapas(df, estado, config, rel) or []
        verificar_solucao_finalizados(df, rel)
        verificar_anexos_locais(links_anexos, config, planilha_path.parent, rel)

        df_pend = linhas_pendentes(df, estado)
        chaves_relevantes["technician_map"] = {chave(v) for v in df_pend.get("Responsável", []) if texto(v)}
        nomes_pend = (
            df_pend["Nome do Solicitante"].fillna("") + " " + df_pend.get("Sobrenome do Solicitante", "").fillna("")
        ).str.strip()
        chaves_relevantes["requester_user_map"] = {chave(v) for v in nomes_pend if v}
    if args.api:
        verificar_api(config, assuntos_sem_map, chaves_relevantes, rel)
    else:
        rel.secao("API do GLPI")
        rel.aviso("Teste de API não solicitado (use --api para validar tokens e IDs no servidor).")

    saida = rel.texto_final()
    print(saida)
    relatorio_path.parent.mkdir(parents=True, exist_ok=True)
    relatorio_path.write_text(saida, encoding="utf-8")
    print(f"\nRelatório salvo em: {relatorio_path}")

    if rel.erros:
        print(f"\nBLOQUEADO: {rel.erros} erro(s) crítico(s) encontrado(s). Corrija antes de continuar.", file=sys.stderr)
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
