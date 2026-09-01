# -*- coding: utf-8 -*-
"""Exportação mensal para validação de culpabilidade do motorista (Código 4).

Fluxo (confirmado com a Qualidade):
    1. A Qualidade classifica uma ocorrência com o Código 4 do Quadro 1
       ("Comportamento ou condição inadequada/insegura (TQUIM)").
    2. Este script gera uma planilha com essas ocorrências para a Qualidade
       mandar aos setores (Frota, Gerência, RH, Programação/Operacional,
       conforme o caso) validarem se foi mesmo culpa do motorista. Isso
       acontece TODO por fora do sistema (e-mail/planilha) — os setores não
       têm acesso ao GLPI para isso.
    3. A Qualidade reúne as respostas dos setores, decide (se houver
       divergência entre setores, a ocorrência NÃO entra na avaliação) e
       devolve a MESMA planilha preenchida na coluna de decisão.
    4. O script `importar_validacao_qualidade.py` lê essa planilha devolvida
       e atualiza o campo "Inserir na avaliação motorista?" de cada chamado.

Importante sobre reexecução mensal: cada chamado só é considerado "resolvido"
depois que a decisão dele foi IMPORTADA de volta (ver
`estado_validacao_qualidade.json`). Enquanto isso não acontece, ele volta a
aparecer nas exportações seguintes (marcado como "já enviado antes"), porque
setores podem demorar a responder. Isso é necessário porque o campo do GLPI
"Inserir na avaliação motorista?" é um yesno que nasce com valor 1 por
padrão MESMO em chamados que a Qualidade nunca tocou — ou seja, não dá pra
usar esse campo pra saber se already foi decidido. Quem decide isso é o
arquivo de estado local, não o GLPI.

Uso:
    python exportar_validacao_qualidade.py [caminho_saida.xlsx]

Requer config.json na mesma pasta.
"""
import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ENTITY_OCORRENCIAS = 6
CONTAINER_DADOS = 1
CONTAINER_CLASSIFICACAO = 6
CAMPO_ID = 2
CAMPO_DATA = 15
CAMPO_TIPO = 7

CODIGO_QUALIDADE_ALVO = 4  # "Comportamento ou condição inadequada/insegura (TQUIM)"

PASTA = Path(__file__).resolve().parent
ARQ_ESTADO = PASTA / "estado_validacao_qualidade.json"

COR_CABECALHO = "FFD966"
COR_TEXTO_CABECALHO = "333333"
COR_LINHA_PAR = "FFF2CC"
COR_BORDA = "D4D9E2"
COR_REENVIO = "FCE4D6"


def _as_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def api_request(base, app_token, method, path, headers=None, data=None):
    url = base + path
    h = {"App-Token": app_token, "Content-Type": "application/json"}
    if headers:
        h.update(headers)
    body = json.dumps(data).encode("utf-8") if data is not None else None
    req = urllib.request.Request(url, data=body, method=method, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Erro na API ({e.code}): {e.read().decode()}") from e


def carregar_estado():
    if ARQ_ESTADO.exists():
        return json.loads(ARQ_ESTADO.read_text(encoding="utf-8"))
    return {"exportados": {}, "importados": {}}


def salvar_estado(estado):
    ARQ_ESTADO.write_text(json.dumps(estado, ensure_ascii=False, indent=2), encoding="utf-8")


def buscar_chamados_base(sreq):
    params = [
        ("criteria[0][field]", "80"),
        ("criteria[0][searchtype]", "equals"),
        ("criteria[0][value]", str(ENTITY_OCORRENCIAS)),
        ("forcedisplay[0]", str(CAMPO_ID)),
        ("forcedisplay[1]", str(CAMPO_TIPO)),
        ("forcedisplay[2]", str(CAMPO_DATA)),
        ("range", "0-9999"),
    ]
    qs = urllib.parse.urlencode(params)
    res = sreq("GET", f"/search/Ticket?{qs}")
    return res.get("data", [])


def buscar_dados_plugin_por_container(sreq, container_id):
    linhas = sreq("GET", "/PluginFieldsTicketdadosdaocorrncia?range=0-3000")
    return {l["items_id"]: l for l in linhas if l.get("plugin_fields_containers_id") == container_id}


def montar_linhas(sreq, estado):
    print("Buscando lista de chamados (rápido, sem campos de plugin)...")
    base = buscar_chamados_base(sreq)
    print(f"{len(base)} chamado(s) na entidade. Buscando dados dos blocos de campos...")

    dados_container1 = buscar_dados_plugin_por_container(sreq, CONTAINER_DADOS)
    dados_container6 = buscar_dados_plugin_por_container(sreq, CONTAINER_CLASSIFICACAO)

    ja_importados = set(estado["importados"].keys())
    ja_exportados = estado["exportados"]

    print("Montando linhas...")
    saida = []
    for row in base:
        tid = _as_int(row.get(str(CAMPO_ID)))
        dados6 = dados_container6.get(tid, {})

        cod_id = _as_int(dados6.get("plugin_fields_codigoqualidadefielddropdowns_id"))
        if cod_id != CODIGO_QUALIDADE_ALVO:
            continue  # só Código 4 (TQUIM) entra nessa validação

        if str(tid) in ja_importados:
            continue  # decisão já aplicada, não precisa mandar de novo

        dados1 = dados_container1.get(tid, {})
        reenvio = str(tid) in ja_exportados

        linha = {
            "ID": tid,
            "Data": str(row.get(str(CAMPO_DATA)) or "")[:10],
            "Motorista": dados1.get("colaboradormotoristafield") or "",
            "Placa": dados1.get("placatraofield") or "",
            "Local": dados1.get("localdaocorrnciafield") or "",
            "Tipo": row.get(str(CAMPO_TIPO)) or "",
            "Descrição da Ocorrência": dados1.get("descriodaocorrnciafield") or "",
            "Confirmado como culpa do motorista? (Sim/Não)": "",
            "_reenvio": ja_exportados.get(str(tid), {}).get("data_exportado") if reenvio else "",
        }
        saida.append(linha)
    return saida


def cabecalhos():
    return ["ID", "Data", "Motorista", "Placa", "Local", "Tipo", "Descrição da Ocorrência",
            "Confirmado como culpa do motorista? (Sim/Não)", "Observação"]


def escrever_planilha(ws, linhas):
    cols = cabecalhos()
    fonte_padrao = Font(name="Arial", size=10)
    fonte_titulo = Font(name="Arial", size=14, bold=True, color=COR_TEXTO_CABECALHO)
    fonte_subtitulo = Font(name="Arial", size=9, italic=True, color="555555")
    fonte_cabecalho = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_CABECALHO)
    borda = Border(*(Side(style="thin", color=COR_BORDA),) * 4)

    n_cols = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    titulo = ws.cell(row=1, column=1, value="VALIDAÇÃO DE CULPABILIDADE DO MOTORISTA — OCORRÊNCIAS CÓDIGO 4")
    titulo.font = fonte_titulo
    titulo.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1, value=(
        f"Gerado em {datetime.now():%d/%m/%Y %H:%M} — {len(linhas)} ocorrência(s). "
        "Preencher a coluna \"Confirmado como culpa do motorista?\" com Sim ou Não e devolver este arquivo."
    ))
    sub.font = fonte_subtitulo
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height
    ws.row_dimensions[2].height = 16

    header_row = 4
    for col_idx, label in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = fonte_cabecalho
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    linhas_ordenadas = sorted(linhas, key=lambda l: (str(l.get("Motorista") or ""), str(l.get("Data") or "")))

    for i, linha in enumerate(linhas_ordenadas):
        row_idx = header_row + 1 + i
        reenvio = bool(linha.get("_reenvio"))
        fill = PatternFill("solid", fgColor=COR_REENVIO) if reenvio else (
            PatternFill("solid", fgColor=COR_LINHA_PAR) if i % 2 == 0 else PatternFill(fill_type=None))
        obs = f"Já enviado em {linha['_reenvio']}, ainda sem resposta" if reenvio else ""
        valores = {**linha, "Observação": obs}
        for col_idx, label in enumerate(cols, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=valores.get(label, ""))
            c.font = fonte_padrao
            c.fill = fill
            c.border = borda
            c.alignment = Alignment(vertical="center", wrap_text=False)

    larguras = {"ID": 8, "Data": 14, "Motorista": 28, "Placa": 14, "Local": 24, "Tipo": 22,
                "Descrição da Ocorrência": 45, "Confirmado como culpa do motorista? (Sim/Não)": 24,
                "Observação": 32}
    for col_idx, label in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(label, 20)
    ws.sheet_view.showGridLines = False
    return linhas_ordenadas


def main():
    cfg = json.loads((PASTA / "config.json").read_text(encoding="utf-8-sig"))

    if len(sys.argv) > 1:
        saida = Path(sys.argv[1])
    else:
        saida = PASTA / f"Validacao_Qualidade_{datetime.now():%Y-%m}.xlsx"

    base_url = cfg["api_url"]
    app_token = cfg["app_token"]
    user_token = cfg["user_token"]
    init = api_request(base_url, app_token, "GET", "/initSession", {"Authorization": "user_token " + user_token})
    session_token = init["session_token"]

    def sreq(method, path, data=None):
        return api_request(base_url, app_token, method, path, {"Session-Token": session_token}, data)

    try:
        estado = carregar_estado()
        linhas = montar_linhas(sreq, estado)
        print(f"{len(linhas)} ocorrência(s) Código 4 pendente(s) de validação.")

        if not linhas:
            print("Nada pra exportar. Nenhum arquivo gerado.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Validação Qualidade"
        linhas_ordenadas = escrever_planilha(ws, linhas)
        wb.save(saida)

        agora = f"{datetime.now():%d/%m/%Y}"
        for linha in linhas_ordenadas:
            tid = str(linha["ID"])
            if tid not in estado["exportados"]:
                estado["exportados"][tid] = {"data_exportado": agora, "arquivo": saida.name}
        salvar_estado(estado)

        print(f"Pronto: {saida}")
    finally:
        sreq("GET", "/killSession")


if __name__ == "__main__":
    main()
