# -*- coding: utf-8 -*-
"""Base do relatório de avaliação de motoristas (TQUIM > Ocorrências).

Lista, por motorista, todas as ocorrências que a Qualidade já classificou
(campo "Código da Ocorrência Quadro 1 - Qualidade" preenchido) E que foram
marcadas para entrar na avaliação ("Inserir na avaliação motorista?" = Sim).

Ainda NÃO calcula desconto de pontos — falta a tabela da Qualidade dizendo
quantos pontos cada código do Quadro 1 tira. Por enquanto é só a listagem
bruta, pra Qualidade revisar, com uma aba de contagem por motorista (só
contagem de ocorrências, sem pontuação ainda).

IMPORTANTE sobre como os dados são buscados: quase todos os campos vêm
direto da tabela do plugin Fields (glpi_plugin_fields_ticketdadosdaocorrncias),
não do /search/Ticket. Só ID/Tipo/Data vêm da busca nativa. Isso é de
propósito: pedir muitas colunas de plugin ao mesmo tempo no /search/Ticket
deixa a consulta pesada demais e o GLPI trava (testado — nem 120s foi
suficiente com ~30 colunas). Buscar a tabela do plugin direto (sem JOIN
pesado) é rápido e confiável.

Uso:
    python exportar_avaliacao_motoristas.py [caminho_saida.xlsx] [ano]

Requer config.json na mesma pasta (mesmo formato dos outros scripts do
projeto): {"api_url": "...", "app_token": "...", "user_token": "..."}
"""
import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ENTITY_OCORRENCIAS = 6
CONTAINER_DADOS = 1
CONTAINER_CLASSIFICACAO = 6
CAMPO_ID = 2
CAMPO_DATA = 15
CAMPO_TIPO = 7

YESNO = {0: "Não", 1: "Sim", None: ""}


def _as_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def _yesno(v):
    return YESNO.get(_as_int(v), v)


def campo(nome):
    return lambda dados: dados.get(nome) or ""


def campo_fmt(nome, fmt):
    return lambda dados: fmt(dados.get(nome))


def dropdown(nome_coluna_fk, tabela_nomes):
    return lambda dados: tabela_nomes.get(str(dados.get(nome_coluna_fk) or 0), "")


# (rótulo, getter(dados_plugin) -> valor). "dados_plugin" é o dicionário da
# linha do plugin Fields (container 1) pra aquele chamado.
COLUNAS_PLUGIN = [
    ("Placa Tração", campo("placatraofield")),
    ("Código Placa Tração", campo("cdfrotatraofield")),
    ("Placa Semi-Reboque", campo("placasemireboquefield")),
    ("Código Placa Semi-Reboque", campo("cdfrotasemireboquefield")),
    ("Motorista", campo("colaboradormotoristafield")),
    ("Cargo/Função", campo("cargofunofield")),
    ("Depto/Setor", campo("deptosetorfield")),
    ("OC/CT-e", campo("occtefield")),
    ("Produto", campo("produtofield")),
    ("Situação da Carga (Granel/Embalado)", None),  # resolvido via dropdown, ver main()
    ("Situação da Carga (Carregado/Vazio)", None),
    ("Origem", campo("origemfield")),
    ("Destino", campo("destinofield")),
    ("Cliente", campo("clientefield")),
    ("Impacto ao Cliente?", campo_fmt("impactoaoclientefield", _yesno)),
    ("Local da ocorrência", campo("localdaocorrnciafield")),
    ("Responsável pela Análise", campo("responsvelpelaanliseaescorretivafield")),
    ("Descrição da Ocorrência", campo("descriodaocorrnciafield")),
    ("Providências já Tomadas", campo("providnciasjtomadafield")),
    ("Outras Observações", campo("outrasobservaefield")),
    ("Motivo", None),
    ("Responsável (Cliente/Motorista)", None),
    ("Plano de ações", campo("planodeaefield")),
    ("Acionamento Suatrans/Pamcary?", campo_fmt("acionamentosuatranspamcaryfield", _yesno)),
    ("Houve vazamento?", campo_fmt("houvevazamentofield", _yesno)),
    ("Quantidade que vazou", campo("quantidadequevazoufield")),
    ("Levantamento de Custos (descritivo)", campo("levantamentodecustofield")),
    ("Custo Total", campo("custototalfield")),
    ("S.A.", campo("safield")),
    # "Inserir na avaliação motorista?" não vem daqui: o campo antigo
    # (container 1) foi desativado e substituído por um novo, dentro do
    # bloco "Classificação" (container 6) - ver getters_extra em
    # montar_linhas(). Mesma razão do Código Qualidade: mover um campo
    # existente de container quebra a busca; um campo novo, criado direto
    # no container certo, não tem esse problema.
]

COLUNAS_LARGURA = {
    "ID": 8, "Tipo": 22, "Motorista": 28, "Data": 16,
    "Código Qualidade (Quadro 1)": 40, "Responsável pela Análise": 24,
    "Descrição da Ocorrência": 40, "Providências já Tomadas": 30,
    "Outras Observações": 30, "Levantamento de Custos (descritivo)": 30,
}

COR_CABECALHO = "FFFF99"
COR_TEXTO_CABECALHO = "333333"
COR_LINHA_PAR = "FFFACD"
COR_BORDA = "D4D9E2"


def api_request(base, app_token, method, path, headers=None, data=None):
    url = base + path
    h = {"App-Token": app_token, "Content-Type": "application/json"}
    if headers:
        h.update(headers)
    body = json.dumps(data).encode("utf-8") if data is not None else None
    req = urllib.request.Request(url, data=body, method=method, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Erro na API ({e.code}): {e.read().decode()}") from e


def buscar_chamados_base(sreq):
    """Só ID, Tipo (categoria) e Data — rápido, sem nenhum campo de plugin."""
    params = [
        ("criteria[0][field]", "80"),
        ("criteria[0][searchtype]", "equals"),
        ("criteria[0][value]", str(ENTITY_OCORRENCIAS)),
        ("forcedisplay[0]", str(CAMPO_ID)),
        ("forcedisplay[1]", str(CAMPO_TIPO)),
        ("forcedisplay[2]", str(CAMPO_DATA)),
        ("range", "0-9999"),
    ]
    qs = urllib.parse.urlencode(params)
    res = sreq("GET", f"/search/Ticket?{qs}")
    return res.get("data", [])


def buscar_dropdown_nomes(sreq, itemtype):
    dados = sreq("GET", f"/{itemtype}?range=0-200")
    return {str(d["id"]): d["name"] for d in dados}


def buscar_dados_plugin_por_container(sreq, container_id):
    """Mapa items_id -> linha inteira do plugin Fields, só pro container
    pedido. Uma chamada só (sem JOIN, rápida) pra todos os chamados."""
    linhas = sreq("GET", "/PluginFieldsTicketdadosdaocorrncia?range=0-3000")
    return {l["items_id"]: l for l in linhas if l.get("plugin_fields_containers_id") == container_id}


def montar_linhas(sreq, ano):
    print("Buscando lista de chamados (rápido, sem campos de plugin)...")
    base = buscar_chamados_base(sreq)
    print(f"{len(base)} chamado(s) na entidade. Buscando dados dos blocos de campos...")

    dados_container1 = buscar_dados_plugin_por_container(sreq, CONTAINER_DADOS)
    dados_container6 = buscar_dados_plugin_por_container(sreq, CONTAINER_CLASSIFICACAO)

    nomes_situacao = buscar_dropdown_nomes(sreq, "PluginFieldsSituaodacargafieldDropdown")
    nomes_carregado = buscar_dropdown_nomes(sreq, "PluginFieldsCarregadovaziofieldDropdown")
    nomes_motivo = buscar_dropdown_nomes(sreq, "PluginFieldsMotivofieldDropdown")
    nomes_respclimot = buscar_dropdown_nomes(sreq, "PluginFieldsResponsvelclientemotoristafieldDropdown")
    nomes_codigo_qualidade = buscar_dropdown_nomes(sreq, "PluginFieldsCodigoqualidadefieldDropdown")

    getters_extra = {
        "Situação da Carga (Granel/Embalado)": dropdown("plugin_fields_situaodacargafielddropdowns_id", nomes_situacao),
        "Situação da Carga (Carregado/Vazio)": dropdown("plugin_fields_carregadovaziofielddropdowns_id", nomes_carregado),
        "Motivo": dropdown("plugin_fields_motivofielddropdowns_id", nomes_motivo),
        "Responsável (Cliente/Motorista)": dropdown("plugin_fields_responsvelclientemotoristafielddropdowns_id", nomes_respclimot),
    }

    print("Montando linhas...")
    saida = []
    for row in base:
        tid = _as_int(row.get(str(CAMPO_ID)))
        dados1 = dados_container1.get(tid, {})
        dados6 = dados_container6.get(tid, {})

        avaliacao = _as_int(dados6.get("avaliacaomotoristafield"))
        if avaliacao != 1:
            continue

        cod_id = dados6.get("plugin_fields_codigoqualidadefielddropdowns_id")
        codigo_qualidade = nomes_codigo_qualidade.get(str(cod_id), "") if cod_id else ""
        if not codigo_qualidade:
            continue  # ainda não classificado pela Qualidade

        data_str = str(row.get(str(CAMPO_DATA)) or "")[:10]
        try:
            if ano and datetime.strptime(data_str, "%Y-%m-%d").year != ano:
                continue
        except ValueError:
            pass

        linha = {
            "ID": tid,
            "Tipo": row.get(str(CAMPO_TIPO)) or "",
            "Data": row.get(str(CAMPO_DATA)) or "",
            "Código Qualidade (Quadro 1)": codigo_qualidade,
            "Inserir na avaliação motorista?": _yesno(avaliacao),
        }
        for label, getter in COLUNAS_PLUGIN:
            if getter is not None:
                linha[label] = getter(dados1) or ""
            else:
                linha[label] = getters_extra[label](dados1) or ""
        saida.append(linha)
    return saida


def cabecalhos():
    return (["ID", "Tipo", "Data"] + [label for label, _ in COLUNAS_PLUGIN]
            + ["Código Qualidade (Quadro 1)", "Inserir na avaliação motorista?"])


def escrever_aba_listagem(ws, linhas):
    cols = cabecalhos()
    fonte_padrao = Font(name="Arial", size=10)
    fonte_titulo = Font(name="Arial", size=14, bold=True, color=COR_TEXTO_CABECALHO)
    fonte_subtitulo = Font(name="Arial", size=9, italic=True, color="555555")
    fonte_cabecalho = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_CABECALHO)
    borda = Border(*(Side(style="thin", color=COR_BORDA),) * 4)

    n_cols = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    titulo = ws.cell(row=1, column=1, value="AVALIAÇÃO DE MOTORISTAS — OCORRÊNCIAS CLASSIFICADAS PELA QUALIDADE")
    titulo.font = fonte_titulo
    titulo.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1, value=(
        f"Gerado em {datetime.now():%d/%m/%Y %H:%M} — {len(linhas)} ocorrência(s). "
        "SEM cálculo de pontos ainda — falta a tabela da Qualidade."
    ))
    sub.font = fonte_subtitulo
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    header_row = 4
    for col_idx, label in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = fonte_cabecalho
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    linhas_ordenadas = sorted(linhas, key=lambda l: (str(l.get("Motorista") or ""), str(l.get("Data") or "")))

    for i, linha in enumerate(linhas_ordenadas):
        row_idx = header_row + 1 + i
        fill = PatternFill("solid", fgColor=COR_LINHA_PAR) if i % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, label in enumerate(cols, start=1):
            valor = linha.get(label, "")
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.font = fonte_padrao
            c.fill = fill
            c.border = borda
            c.alignment = Alignment(vertical="center", wrap_text=False)

    for col_idx, label in enumerate(cols, start=1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = COLUNAS_LARGURA.get(label, 20)
    ws.sheet_view.showGridLines = False
    return linhas_ordenadas


def escrever_aba_resumo(ws, linhas):
    fonte_padrao = Font(name="Arial", size=10)
    fonte_titulo = Font(name="Arial", size=14, bold=True, color=COR_TEXTO_CABECALHO)
    fonte_cabecalho = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_CABECALHO)
    borda = Border(*(Side(style="thin", color=COR_BORDA),) * 4)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    titulo = ws.cell(row=1, column=1, value="RESUMO POR MOTORISTA (contagem, sem pontuação)")
    titulo.font = fonte_titulo
    titulo.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_row = 3
    for col_idx, label in enumerate(["Motorista", "Qtde. de Ocorrências"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = fonte_cabecalho
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    contagem = Counter(str(linha.get("Motorista") or "(sem motorista informado)") for linha in linhas)
    for i, (motorista, qtde) in enumerate(sorted(contagem.items(), key=lambda x: -x[1])):
        row_idx = header_row + 1 + i
        fill = PatternFill("solid", fgColor=COR_LINHA_PAR) if i % 2 == 0 else PatternFill(fill_type=None)
        c1 = ws.cell(row=row_idx, column=1, value=motorista)
        c2 = ws.cell(row=row_idx, column=2, value=qtde)
        for c in (c1, c2):
            c.font = fonte_padrao
            c.fill = fill
            c.border = borda
        c2.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.sheet_view.showGridLines = False


def main():
    pasta = Path(__file__).resolve().parent
    cfg = json.loads((pasta / "config.json").read_text(encoding="utf-8-sig"))

    ano = int(sys.argv[2]) if len(sys.argv) > 2 else datetime.now().year
    if len(sys.argv) > 1:
        saida = Path(sys.argv[1])
    else:
        saida = pasta / f"Avaliacao_Motoristas_{ano}.xlsx"

    base_url = cfg["api_url"]
    app_token = cfg["app_token"]
    user_token = cfg["user_token"]
    init = api_request(base_url, app_token, "GET", "/initSession", {"Authorization": "user_token " + user_token})
    session_token = init["session_token"]

    def sreq(method, path, data=None):
        return api_request(base_url, app_token, method, path, {"Session-Token": session_token}, data)

    try:
        linhas = montar_linhas(sreq, ano)
        print(f"{len(linhas)} ocorrência(s) classificada(s) e marcada(s) pra avaliação em {ano}.")

        wb = Workbook()
        wb.remove(wb.active)
        ws_lista = wb.create_sheet("Ocorrências Classificadas")
        linhas_ordenadas = escrever_aba_listagem(ws_lista, linhas)
        ws_resumo = wb.create_sheet("Resumo por Motorista")
        escrever_aba_resumo(ws_resumo, linhas_ordenadas)
        wb.save(saida)
        print(f"Pronto: {saida}")
    finally:
        sreq("GET", "/killSession")


if __name__ == "__main__":
    main()
