# -*- coding: utf-8 -*-
"""Exporta um relatório em Excel de todos os chamados da entidade TQUIM > Ocorrências
do ano corrente, com os campos nativos do chamado e os campos personalizados do
plugin Fields.

O arquivo sai com três partes:
  - "Anual <ano>": painel do ano — números do período, consulta interativa por
    mês e por tipo (listas suspensas), quebra por mês e por tipo com gráficos, e
    as instruções de como pedir o relatório por e-mail;
  - uma aba por mês que tiver pelo menos um chamado, com a listagem detalhada;
  - "Dados" (oculta): a listagem do ano inteiro, que é a base somada pelo painel.

Uso:
    python exportar_relatorio_ocorrencias.py [caminho_saida.xlsx] [ano]

Requer config.json na mesma pasta (mesmo formato usado nos outros scripts do projeto):
    {"api_url": "...", "app_token": "...", "user_token": "..."}
"""
import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ENTITY_OCORRENCIAS = 6
CAMPO_DATA_ABERTURA = 15  # search-option usado tanto pra coluna quanto pra agrupar por mês
CAMPO_ID = 2
CAMPO_TIPO = 7

# Os campos do plugin Fields vêm da tabela do plugin, não do /search/Ticket
# (ver o comentário em buscar_chamados). Os dois blocos dividem a mesma tabela,
# separados pelo container.
PLUGIN_ITEMTYPE = "PluginFieldsTicketdadosdaocorrncia"
CONTAINER_DADOS = 1           # "Dados da Ocorrência" (inline na abertura)
CONTAINER_CLASSIFICACAO = 6   # "Classificação" (aba, preenchida pela Qualidade)

DROPDOWNS = [
    "PluginFieldsSituaodacargafieldDropdown",
    "PluginFieldsCarregadovaziofieldDropdown",
    "PluginFieldsMotivofieldDropdown",
    "PluginFieldsResponsvelclientemotoristafieldDropdown",
    "PluginFieldsClassificaofieldDropdown",
]

MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

YESNO = {0: "Não", 1: "Sim", None: ""}


def _as_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def _yesno(v):
    return YESNO.get(_as_int(v), v)


def _num(v):
    """O GLPI devolve o custo como texto; vira número pra poder somar no indicador."""
    if v is None or str(v).strip() == "":
        return ""
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return v          # não era número: mantém o que veio


# Colunas do relatório, na ordem e com os nomes definidos pela TQUIM.
# Cada linha é montada em montar_linha(), que já resolve de onde vem cada valor
# (busca nativa, tabela do plugin, lista suspensa ou anexo).
COLUNAS = [
    "ID",
    "Tipo",
    "Data",
    "Placa Tração",
    "Código Placa Tração",
    "Placa Semi-Reboque",
    "Código Placa Semi-Reboque",
    "Colaborador",
    "Cargo/Função",
    "Depto/Setor",
    "Inserir na avaliação motorista?",
    "OC/CT-e",
    "Produto",
    "Situação da Carga (Granel/Embalado/Carregado/Vazio)",
    "Origem",
    "Destino",
    "Cliente",
    "Impacto ao Cliente?",
    "Local da ocorrência",
    "Responsável pela Análise",
    "Descrição da Ocorrência",
    "Ações Imediatas",
    "Observações e comentários",
    "Descrição da Jornada",
    "Motivo",
    "Responsável (Cliente/Motorista)",
    "Plano de ações",
    "Acionamento Suatrans/Pamcary?",
    "Houve vazamento?",
    "Quantidade que vazou",
    "Fotos",
    "Levantamento de Custos (descritivo)",
    "Custo Total",
    "S.A.",
    "Classificação",  # deixado em branco/editável para a Qualidade preencher
]

# Cores do modelo de abertura de Ocorrência (mesmas do e-mail de notificação)
COR_CABECALHO = "FFFF99"
COR_TEXTO_CABECALHO = "333333"
COR_LINHA_PAR = "FFFACD"
COR_BORDA = "D4D9E2"

# Cores do painel anual (aba do indicador)
COR_INDICADOR = "1F4E79"
COR_INTERATIVO = "1E7A5F"
COR_PEDIDO = "C0622A"

# Como pedir o relatório fora do dia agendado — mesmo gatilho aceito pelo
# verificar_solicitacao_relatorio.php, que lê os chamados abertos pelo coletor.
EMAIL_PEDIDO = "chamados.ti@tquim.com.br"
PALAVRA_CHAVE_PEDIDO = "RELATORIO OCORRENCIAS"

LARGURAS = {
    "ID": 8, "Tipo": 22, "Data": 16, "Descrição da Ocorrência": 40,
    "Ações Imediatas": 30, "Observações e comentários": 30,
    "Descrição da Jornada": 34, "Fotos": 30,
    "Levantamento de Custos (descritivo)": 30,
}


def api_request(base, app_token, method, path, headers=None, data=None):
    url = base + path
    h = {"App-Token": app_token, "Content-Type": "application/json"}
    if headers:
        h.update(headers)
    body = json.dumps(data).encode("utf-8") if data is not None else None
    req = urllib.request.Request(url, data=body, method=method, headers=h)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Erro na API ({e.code}): {e.read().decode()}") from e


def buscar_chamados(app_ctx):
    """Busca só ID, Tipo e Data pela busca nativa.

    Pedir as ~33 colunas do plugin de uma vez no /search/Ticket faz a busca
    nativa parar de responder: medido neste GLPI, 15 colunas levam 15s, 22
    colunas passam de 2 minutos e 33 colunas não terminam nem em 15 minutos
    (e em algumas combinações o JOIN ainda multiplica as linhas — 434 chamados
    voltando como 998). Os campos do plugin vêm da tabela do plugin, numa
    chamada só, como o exportar_avaliacao_motoristas.py já fazia.
    """
    base, app_token, sreq = app_ctx
    params = [
        ("criteria[0][field]", "80"),
        ("criteria[0][searchtype]", "equals"),
        ("criteria[0][value]", str(ENTITY_OCORRENCIAS)),
        ("range", "0-9999"),
    ]
    for col_id in (CAMPO_ID, CAMPO_TIPO, CAMPO_DATA_ABERTURA):
        params.append(("forcedisplay[]", str(col_id)))
    qs = urllib.parse.urlencode(params)
    res = sreq("GET", f"/search/Ticket?{qs}")

    vistos = {}
    for linha in res.get("data", []):
        vistos.setdefault(str(linha.get(str(CAMPO_ID))), linha)
    return list(vistos.values())


def buscar_dropdown(app_ctx, itemtype):
    """id -> nome de uma lista suspensa do plugin."""
    base, app_token, sreq = app_ctx
    try:
        dados = sreq("GET", f"/{itemtype}?range=0-500")
    except RuntimeError:
        return {}
    return {str(d.get("id")): d.get("name") or "" for d in dados}


def buscar_dados_plugin(app_ctx):
    """Lê a tabela do plugin de uma vez e indexa por chamado.

    Os dois blocos (Dados da Ocorrência e Classificação) dividem a mesma tabela
    física, separados por plugin_fields_containers_id — por isso o retorno é um
    dicionário por container.
    """
    base, app_token, sreq = app_ctx
    linhas = sreq("GET", f"/{PLUGIN_ITEMTYPE}?range=0-9999")
    por_container = {CONTAINER_DADOS: {}, CONTAINER_CLASSIFICACAO: {}}
    for linha in linhas:
        if linha.get("itemtype") != "Ticket":
            continue
        cid = linha.get("plugin_fields_containers_id")
        if cid in por_container:
            por_container[cid][str(linha.get("items_id"))] = linha
    return por_container


def buscar_anexos_todos(app_ctx):
    """chamado -> nomes dos anexos, em duas chamadas em vez de uma por chamado.

    Uma chamada /Ticket/<id>/Document_Item por chamado levava ~0,7s, o que dava
    mais de 5 minutos só nessa etapa.
    """
    base, app_token, sreq = app_ctx
    try:
        vinculos = sreq("GET", "/Document_Item?range=0-9999")
        documentos = sreq("GET", "/Document?range=0-9999")
    except RuntimeError:
        return {}

    nome_por_doc = {}
    for d in documentos:
        nome_por_doc[str(d.get("id"))] = d.get("filename") or d.get("name") or f"documento-{d.get('id')}"

    por_ticket = {}
    for v in vinculos:
        if v.get("itemtype") != "Ticket":
            continue
        doc_id = str(v.get("documents_id"))
        if doc_id in nome_por_doc:
            por_ticket.setdefault(str(v.get("items_id")), []).append(nome_por_doc[doc_id])
    return {k: ", ".join(v) for k, v in por_ticket.items()}


def montar_linha(chamado, dados, classificacao, anexos, nomes):
    """Junta o que veio da busca nativa, da tabela do plugin e dos anexos
    numa única linha, já com os nomes das listas suspensas resolvidos."""
    dados = dados or {}
    classificacao = classificacao or {}

    def dd(coluna_fk, itemtype):
        return nomes.get(itemtype, {}).get(str(dados.get(coluna_fk)), "")

    situacao = " / ".join(p for p in (
        dd("plugin_fields_situaodacargafielddropdowns_id", "PluginFieldsSituaodacargafieldDropdown"),
        dd("plugin_fields_carregadovaziofielddropdowns_id", "PluginFieldsCarregadovaziofieldDropdown"),
    ) if p)

    # o campo da avaliação foi movido pro bloco Classificação; quando o chamado
    # ainda não tem esse bloco preenchido, cai no campo antigo do bloco Dados
    avaliacao = classificacao.get("avaliacaomotoristafield")
    if avaliacao is None:
        avaliacao = dados.get("inserirnaavaliaomotoristafield")

    return {
        "ID": chamado.get(str(CAMPO_ID)),
        "Tipo": chamado.get(str(CAMPO_TIPO)),
        "Data": chamado.get(str(CAMPO_DATA_ABERTURA)),
        "Placa Tração": dados.get("placatraofield"),
        "Código Placa Tração": dados.get("cdfrotatraofield"),
        "Placa Semi-Reboque": dados.get("placasemireboquefield"),
        "Código Placa Semi-Reboque": dados.get("cdfrotasemireboquefield"),
        "Colaborador": dados.get("colaboradormotoristafield"),
        "Cargo/Função": dados.get("cargofunofield"),
        "Depto/Setor": dados.get("deptosetorfield"),
        "Inserir na avaliação motorista?": _yesno(avaliacao),
        "OC/CT-e": dados.get("occtefield"),
        "Produto": dados.get("produtofield"),
        "Situação da Carga (Granel/Embalado/Carregado/Vazio)": situacao,
        "Origem": dados.get("origemfield"),
        "Destino": dados.get("destinofield"),
        "Cliente": dados.get("clientefield"),
        "Impacto ao Cliente?": _yesno(dados.get("impactoaoclientefield")),
        "Local da ocorrência": dados.get("localdaocorrnciafield"),
        "Responsável pela Análise": dados.get("responsvelpelaanliseaescorretivafield"),
        "Descrição da Ocorrência": dados.get("descriodaocorrnciafield"),
        "Ações Imediatas": dados.get("providnciasjtomadafield"),
        "Observações e comentários": dados.get("outrasobservaefield"),
        "Descrição da Jornada": _jornada_plugin(dados),
        "Motivo": dd("plugin_fields_motivofielddropdowns_id", "PluginFieldsMotivofieldDropdown"),
        "Responsável (Cliente/Motorista)": dd(
            "plugin_fields_responsvelclientemotoristafielddropdowns_id",
            "PluginFieldsResponsvelclientemotoristafieldDropdown"),
        "Plano de ações": dados.get("planodeaefield"),
        "Acionamento Suatrans/Pamcary?": _yesno(dados.get("acionamentosuatranspamcaryfield")),
        "Houve vazamento?": _yesno(dados.get("houvevazamentofield")),
        "Quantidade que vazou": dados.get("quantidadequevazoufield"),
        "Fotos": anexos,
        "Levantamento de Custos (descritivo)": dados.get("levantamentodecustofield"),
        "Custo Total": _num(dados.get("custototalfield")),
        "S.A.": dados.get("safield"),
        "Classificação": dd("plugin_fields_classificaofielddropdowns_id",
                            "PluginFieldsClassificaofieldDropdown"),
    }


def _jornada_plugin(dados):
    partes = []
    if dados.get("inciodajornadafield") or dados.get("fimdajornadafield"):
        partes.append(f"Jornada: {dados.get('inciodajornadafield') or '-'} "
                      f"até {dados.get('fimdajornadafield') or '-'}")
    if dados.get("inciodocarregamentofield") or dados.get("fimdocarregamentofield"):
        partes.append(f"Carregamento: {dados.get('inciodocarregamentofield') or '-'} "
                      f"até {dados.get('fimdocarregamentofield') or '-'}")
    return " | ".join(partes)


def mes_da_linha(linha):
    valor = linha.get("Data")
    if not valor:
        return None
    try:
        return datetime.strptime(str(valor)[:10], "%Y-%m-%d")
    except ValueError:
        return None


def filtrar_por_ano(linhas, ano):
    return [l for l in linhas if (d := mes_da_linha(l)) and d.year == ano]


def agrupar_por_mes(linhas):
    grupos = {}
    for linha in linhas:
        data = mes_da_linha(linha)
        if not data:
            continue
        grupos.setdefault(data.month, []).append(linha)
    return grupos


def escrever_aba(ws, linhas):
    fonte_padrao = Font(name="Arial", size=10)
    fonte_titulo = Font(name="Arial", size=14, bold=True, color=COR_TEXTO_CABECALHO)
    fonte_subtitulo = Font(name="Arial", size=9, italic=True, color="555555")
    fonte_cabecalho = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_CABECALHO)
    borda = Border(*(Side(style="thin", color=COR_BORDA),) * 4)

    n_cols = len(COLUNAS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    # "Julho - 2026" -> "Ocorrências - Julho 2026"
    titulo = ws.cell(row=1, column=1, value=f"Ocorrências - {ws.title.replace(' - ', ' ')}")
    titulo.font = fonte_titulo
    titulo.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1, value=f"Gerado em {datetime.now():%d/%m/%Y %H:%M} — {len(linhas)} chamado(s)")
    sub.font = fonte_subtitulo
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    header_row = 4
    for col_idx, label in enumerate(COLUNAS, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = fonte_cabecalho
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for i, linha in enumerate(linhas):
        row_idx = header_row + 1 + i
        fill = PatternFill("solid", fgColor=COR_LINHA_PAR) if i % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, label in enumerate(COLUNAS, start=1):
            valor = linha.get(label)
            if valor is None:
                valor = ""
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.font = fonte_padrao
            c.fill = fill
            c.border = borda
            c.alignment = Alignment(vertical="center", wrap_text=False)

    for col_idx, label in enumerate(COLUNAS, start=1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = LARGURAS.get(label, 20)
    ws.sheet_view.showGridLines = False

    # Coluna "Classificação" fica com o campo desativado no GLPI (preenchida
    # pela Qualidade direto aqui na planilha) — deixa pronta pra receber uma
    # lista suspensa assim que os valores forem definidos.
    classificacao_col = COLUNAS.index("Classificação") + 1
    letra = get_column_letter(classificacao_col)
    if len(linhas) > 0:
        dv = DataValidation(type="list", formula1='"Leve,Moderada,Grave,Crítica"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letra}{header_row + 1}:{letra}{header_row + len(linhas)}")


def escrever_dados(ws, linhas):
    """Listagem crua do ano, numa aba oculta: é a base que o indicador soma.
    Cabeçalho na linha 1 e dados a partir da 2, pra manter as faixas previsíveis."""
    ws.append(list(COLUNAS))
    for c in ws[1]:
        c.font = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_CABECALHO)
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)

    for linha in linhas:
        ws.append(["" if linha.get(label) is None else linha.get(label) for label in COLUNAS])

    col_custo = COLUNAS.index("Custo Total") + 1
    for r in range(2, len(linhas) + 2):
        ws.cell(row=r, column=col_custo).number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    if linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{len(linhas) + 1}"


def escrever_indicador(ws, linhas, ano):
    """Painel do ano: números do período, consulta interativa por mês/tipo,
    quebra por mês e por tipo, e como pedir o relatório por e-mail."""
    ws.sheet_view.showGridLines = False

    f_titulo = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    f_sub = Font(name="Arial", size=9, italic=True, color="555555")
    f_secao = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    f_rotulo = Font(name="Arial", size=9, bold=True, color="444444")
    f_txt = Font(name="Arial", size=10)
    f_bold = Font(name="Arial", size=10, bold=True)
    borda = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

    for col, larg in zip("ABCDEFGH", [22, 14, 14, 4, 26, 14, 14, 14]):
        ws.column_dimensions[col].width = larg
    for col in ("T", "U", "V"):
        ws.column_dimensions[col].hidden = True

    def faixa(row, texto, cor):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=texto)
        c.font = f_secao
        c.fill = PatternFill("solid", fgColor=cor)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def bloco(row_rot, row_val, itens, cor_fundo, cor_valor, tamanho):
        for col, rotulo, formula, fmt in itens:
            col2 = get_column_letter(column_index_from_string(col) + 1)
            ws.merge_cells(f"{col}{row_rot}:{col2}{row_rot}")
            ws.merge_cells(f"{col}{row_val}:{col2}{row_val}")
            r = ws[f"{col}{row_rot}"]
            r.value, r.font = rotulo, f_rotulo
            r.alignment = Alignment(horizontal="center")
            r.fill = PatternFill("solid", fgColor=cor_fundo)
            v = ws[f"{col}{row_val}"]
            v.value = formula
            v.font = Font(name="Arial", size=tamanho, bold=True, color=cor_valor)
            v.number_format = fmt
            v.alignment = Alignment(horizontal="center")
            v.fill = PatternFill("solid", fgColor=cor_fundo)

    def coluna(label):
        return get_column_letter(COLUNAS.index(label) + 1)

    fim = len(linhas) + 1
    faixa_de = lambda label: f"Dados!${coluna(label)}$2:${coluna(label)}${fim}"
    R_TIPO, R_DATA = faixa_de("Tipo"), faixa_de("Data")
    R_IMP, R_AVAL = faixa_de("Impacto ao Cliente?"), faixa_de("Inserir na avaliação motorista?")
    R_CUSTO = faixa_de("Custo Total")

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Ocorrências - Anual {ano}"
    t.font = f_titulo
    t.fill = PatternFill("solid", fgColor=COR_INDICADOR)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Gerado em {datetime.now():%d/%m/%Y %H:%M} — {len(linhas)} chamado(s)"
    ws["A2"].font = f_sub
    ws["A2"].alignment = Alignment(horizontal="center")

    faixa(4, "PANORAMA DO ANO", COR_INDICADOR)
    bloco(5, 6, [
        ("A", "Total de ocorrências", f"=COUNTA({R_TIPO})", "0"),
        ("C", "Meses com registro", "=SUMPRODUCT((B17:B28>0)*1)", "0"),
        ("E", "Com impacto ao cliente", f'=COUNTIF({R_IMP},"Sim")', "0"),
        ("G", "Custo total lançado (R$)", f"=SUM({R_CUSTO})", "#,##0"),
    ], "F2F2F2", COR_INDICADOR, 20)
    ws.row_dimensions[6].height = 30

    ws.merge_cells("A7:H7")
    ws["A7"] = ('O custo considera apenas os registros com "Custo Total" preenchido. A listagem '
                "completa está na aba Dados (oculta: botão direito em uma aba > Reexibir).")
    ws["A7"].font = f_sub

    faixa(9, "CONSULTA INTERATIVA  —  escolha nas listas abaixo", COR_INTERATIVO)
    ws["A10"], ws["E10"] = "Mês:", "Tipo:"
    ws["A10"].font = ws["E10"].font = f_bold
    ws.merge_cells("F10:H10")
    ws["B10"], ws["F10"] = "(todos)", "(todos)"
    for cel in ("B10", "F10"):
        ws[cel].font = Font(name="Arial", size=10, bold=True, color="C00000")
        ws[cel].fill = PatternFill("solid", fgColor="FFF2CC")
        ws[cel].border = borda
        ws[cel].alignment = Alignment(horizontal="center")

    meses_presentes = sorted({str(l.get("Data"))[:7] for l in linhas if l.get("Data")})
    tipos = sorted({str(l.get("Tipo")).strip() for l in linhas if l.get("Tipo")})

    for i, m in enumerate(meses_presentes, start=2):
        ws[f"T{i}"] = m
    ws["U2"] = "(todos)"
    for i, m in enumerate(meses_presentes, start=3):
        ws[f"U{i}"] = MESES[int(m[5:7])]
    ws["V2"] = "(todos)"
    for i, tp in enumerate(tipos, start=3):
        ws[f"V{i}"] = tp

    dv_mes = DataValidation(type="list", formula1=f"=$U$2:$U${len(meses_presentes) + 2}", allow_blank=False)
    ws.add_data_validation(dv_mes)
    dv_mes.add("B10")
    dv_tipo = DataValidation(type="list", formula1=f"=$V$2:$V${len(tipos) + 2}", allow_blank=False)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add("F10")

    # "(todos)" vira o curinga "*"; um mês vira o prefixo "AAAA-MM*" da data
    ws["T20"] = (f'=IF($B$10="(todos)","*",INDEX($T$2:$T${len(meses_presentes) + 1},'
                 f'MATCH($B$10,$U$3:$U${len(meses_presentes) + 2},0))&"*")')
    ws["T21"] = '=IF($F$10="(todos)","*",$F$10)'

    bloco(12, 13, [
        ("A", "Ocorrências no filtro", f"=COUNTIFS({R_DATA},$T$20,{R_TIPO},$T$21)", "0"),
        ("C", "Com impacto ao cliente", f'=COUNTIFS({R_DATA},$T$20,{R_TIPO},$T$21,{R_IMP},"Sim")', "0"),
        ("E", "Na avaliação do motorista", f'=COUNTIFS({R_DATA},$T$20,{R_TIPO},$T$21,{R_AVAL},"Sim")', "0"),
        ("G", "Custo no filtro (R$)", f"=SUMIFS({R_CUSTO},{R_DATA},$T$20,{R_TIPO},$T$21)", "#,##0"),
    ], "DCE6F1", COR_INTERATIVO, 18)
    ws.row_dimensions[13].height = 28

    faixa(15, "OCORRÊNCIAS POR MÊS", COR_INDICADOR)
    for cel, txt in (("A16", "Mês"), ("B16", "Qtde")):
        ws[cel] = txt
        ws[cel].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws[cel].fill = PatternFill("solid", fgColor="7F7F7F")
        ws[cel].alignment = Alignment(horizontal="center")
    for i, mes in enumerate(range(1, 13)):
        lin = 17 + i
        ws[f"A{lin}"] = MESES[mes]
        ws[f"B{lin}"] = f'=COUNTIF({R_DATA},"{ano}-{mes:02d}*")'
        for cel in (f"A{lin}", f"B{lin}"):
            ws[cel].font = f_txt
            ws[cel].border = borda
        ws[f"B{lin}"].alignment = Alignment(horizontal="center")
    ws.conditional_formatting.add(
        "B17:B28", DataBarRule(start_type="num", start_value=0, end_type="max", color=COR_INDICADOR))

    g_mes = BarChart()
    g_mes.type, g_mes.legend = "col", None
    g_mes.title = "Ocorrências por mês"
    g_mes.height, g_mes.width = 7.5, 13
    g_mes.add_data(Reference(ws, min_col=2, min_row=16, max_row=28), titles_from_data=True)
    g_mes.set_categories(Reference(ws, min_col=1, min_row=17, max_row=28))
    ws.add_chart(g_mes, "E16")

    faixa(30, "OCORRÊNCIAS POR TIPO", COR_INDICADOR)
    for cel, txt in (("A31", "Tipo"), ("B31", "Qtde")):
        ws[cel] = txt
        ws[cel].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws[cel].fill = PatternFill("solid", fgColor="7F7F7F")
        ws[cel].alignment = Alignment(horizontal="center")

    contagem = Counter(str(l.get("Tipo")).strip() for l in linhas if l.get("Tipo"))
    top = [t for t, _ in contagem.most_common(10)]
    for i, tp in enumerate(top):
        lin = 32 + i
        ws[f"A{lin}"] = tp
        ws[f"A{lin}"].alignment = Alignment(wrap_text=True)
        ws[f"B{lin}"] = f"=COUNTIF({R_TIPO},A{lin})"
        ws[f"B{lin}"].alignment = Alignment(horizontal="center")
        for cel in (f"A{lin}", f"B{lin}"):
            ws[cel].font = f_txt
            ws[cel].border = borda
    fim_tipo = 31 + len(top)
    lin_demais = fim_tipo + 1
    ws[f"A{lin_demais}"] = "Demais tipos"
    ws[f"B{lin_demais}"] = f"=COUNTA({R_TIPO})-SUM(B32:B{fim_tipo})"
    for cel in (f"A{lin_demais}", f"B{lin_demais}"):
        ws[cel].font = Font(name="Arial", size=10, italic=True)
        ws[cel].border = borda
    ws[f"B{lin_demais}"].alignment = Alignment(horizontal="center")
    ws.conditional_formatting.add(
        f"B32:B{lin_demais}", DataBarRule(start_type="num", start_value=0, end_type="max", color=COR_PEDIDO))

    if top:
        g_tipo = BarChart()
        g_tipo.type, g_tipo.legend = "bar", None
        g_tipo.title = "Top 10 tipos de ocorrência"
        g_tipo.height, g_tipo.width = 9, 13
        g_tipo.add_data(Reference(ws, min_col=2, min_row=31, max_row=fim_tipo), titles_from_data=True)
        g_tipo.set_categories(Reference(ws, min_col=1, min_row=32, max_row=fim_tipo))
        ws.add_chart(g_tipo, "E31")

    rodape = lin_demais + 2
    faixa(rodape, "COMO RECEBER ESTE RELATÓRIO POR E-MAIL", COR_PEDIDO)
    instrucoes = [
        (f"1. Envie um e-mail para {EMAIL_PEDIDO}", True),
        (f"2. Escreva no ASSUNTO: {PALAVRA_CHAVE_PEDIDO}", True),
        ("3. O relatório do ano corrente chega como anexo, em resposta ao próprio e-mail.", False),
        ("", False),
        ("O pedido precisa partir de um endereço autorizado. O corpo do e-mail pode ir vazio —", False),
        ("o que vale é a palavra-chave no assunto. Fora isso, o relatório também é enviado", False),
        ("automaticamente todo dia 1º de cada mês.", False),
    ]
    for i, (texto, forte) in enumerate(instrucoes):
        lin = rodape + 1 + i
        ws.merge_cells(start_row=lin, start_column=1, end_row=lin, end_column=8)
        c = ws.cell(row=lin, column=1, value=texto)
        c.font = Font(name="Arial", size=10, bold=forte, color=COR_INDICADOR if forte else "444444")
        c.alignment = Alignment(indent=1)


def montar_planilha(linhas_ano, ano, caminho_saida):
    wb = Workbook()
    wb.remove(wb.active)

    ws_anual = wb.create_sheet(f"Anual {ano}")

    grupos = agrupar_por_mes(linhas_ano)
    for mes in sorted(grupos):
        ws_mes = wb.create_sheet(f"{MESES[mes]} - {ano}")
        escrever_aba(ws_mes, grupos[mes])

    ws_dados = wb.create_sheet("Dados")
    escrever_dados(ws_dados, linhas_ano)
    ws_dados.sheet_state = "hidden"

    escrever_indicador(ws_anual, linhas_ano, ano)

    wb.save(caminho_saida)


def main():
    pasta = Path(__file__).resolve().parent
    cfg = json.loads((pasta / "config.json").read_text(encoding="utf-8-sig"))

    ano = int(sys.argv[2]) if len(sys.argv) > 2 else datetime.now().year
    if len(sys.argv) > 1:
        saida = Path(sys.argv[1])
    else:
        saida = pasta / f"Relatorio_Ocorrencias_{ano}.xlsx"

    base = cfg["api_url"]
    app_token = cfg["app_token"]
    user_token = cfg["user_token"]
    init = api_request(base, app_token, "GET", "/initSession", {"Authorization": "user_token " + user_token})
    session_token = init["session_token"]

    def sreq(method, path, data=None):
        return api_request(base, app_token, method, path, {"Session-Token": session_token}, data)

    app_ctx = (base, app_token, sreq)

    try:
        print("Buscando chamados da entidade Ocorrências...")
        chamados = buscar_chamados(app_ctx)
        print(f"{len(chamados)} chamado(s) na entidade. Lendo os campos do plugin...")
        plugin = buscar_dados_plugin(app_ctx)

        print("Lendo as listas suspensas...")
        nomes = {it: buscar_dropdown(app_ctx, it) for it in DROPDOWNS}

        print("Lendo os anexos...")
        anexos = buscar_anexos_todos(app_ctx)

        print("Montando as linhas...")
        linhas = []
        for chamado in chamados:
            tid = str(chamado.get(str(CAMPO_ID)))
            linhas.append(montar_linha(
                chamado,
                plugin[CONTAINER_DADOS].get(tid),
                plugin[CONTAINER_CLASSIFICACAO].get(tid),
                anexos.get(tid, ""),
                nomes,
            ))

        linhas_ano = filtrar_por_ano(linhas, ano)
        print(f"{len(linhas_ano)} chamado(s) de {ano}. Montando planilha...")
        montar_planilha(linhas_ano, ano, saida)
        print(f"Pronto: {saida}")
    finally:
        sreq("GET", "/killSession")


if __name__ == "__main__":
    main()
