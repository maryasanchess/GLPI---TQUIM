# -*- coding: utf-8 -*-
"""Lê a planilha de validação (devolvida pela Qualidade, já com as respostas
dos setores) e atualiza no GLPI o campo "Inserir na avaliação motorista?" de
cada chamado.

Suporta o fluxo real: a Qualidade tira o relatório de Código 4 e separa em
planilhas menores por tipo de ocorrência, uma pra cada setor (ex.: Jornada
de Trabalho vai pra RH e Rastreamento). Cada setor devolve a sua parte
separada, e a Qualidade anexa uma de cada vez. Como o arquivo anexado pode
ser só um recorte do relatório original, cada linha é conferida contra o
GLPI por ID **e** Data antes de aplicar qualquer mudança — se a data da
planilha não bater com a data real do chamado no GLPI, a linha é pulada e
avisada, em vez de arriscar atualizar o chamado errado.

O que este script NÃO faz (de propósito, por decisão da Qualidade): não
calcula desconto de pontos, não decide o que é maioria/empate entre setores.
Isso é feito por fora, pela Qualidade — o sistema só reflete a decisão final
que já veio pronta na planilha.

Uso:
    python importar_validacao_qualidade.py caminho_planilha_preenchida.xlsx

Requer config.json na mesma pasta. Não dispara notificação nenhuma (mesmo
padrão usado no importador geral: _disablenotifications + _disablenotif).
"""
import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ENTITY_OCORRENCIAS = 6
CONTAINER_CLASSIFICACAO = 6
CAMPO_ID = 2
CAMPO_DATA = 15
PASTA = Path(__file__).resolve().parent
ARQ_ESTADO = PASTA / "estado_validacao_qualidade.json"

RESPOSTAS_SIM = {"sim", "s", "yes", "1"}
RESPOSTAS_NAO = {"não", "nao", "n", "no", "0"}


def _normalizar_data(v):
    """Aceita tanto texto (ISO, como o export gera) quanto data/datetime
    (caso alguém reformate a célula no Excel) e devolve sempre AAAA-MM-DD."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-":
        return s[:10]
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10]


def api_request(base, app_token, method, path, headers=None, data=None):
    url = base + path
    h = {"App-Token": app_token, "Content-Type": "application/json"}
    if headers:
        h.update(headers)
    body = json.dumps(data).encode("utf-8") if data is not None else None
    req = urllib.request.Request(url, data=body, method=method, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Erro na API ({e.code}): {e.read().decode()}") from e


def carregar_estado():
    if ARQ_ESTADO.exists():
        return json.loads(ARQ_ESTADO.read_text(encoding="utf-8"))
    return {"exportados": {}, "importados": {}}


def salvar_estado(estado):
    ARQ_ESTADO.write_text(json.dumps(estado, ensure_ascii=False, indent=2), encoding="utf-8")


def ler_planilha(caminho):
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, 8):
        valores = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "ID" in valores:
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("Não encontrei a linha de cabeçalho (coluna \"ID\") na planilha.")

    col_por_nome = {}
    for c in range(1, ws.max_column + 1):
        nome = ws.cell(row=header_row, column=c).value
        if nome:
            col_por_nome[str(nome).strip()] = c

    col_id = col_por_nome.get("ID")
    col_data = col_por_nome.get("Data")
    col_decisao = col_por_nome.get("Confirmado como culpa do motorista? (Sim/Não)")
    if not col_id or not col_data or not col_decisao:
        raise RuntimeError(
            "Planilha não tem as colunas esperadas (\"ID\", \"Data\" e a coluna de decisão). "
            "Use a planilha original gerada pelo exportar_validacao_qualidade.py (pode estar "
            "recortada em menos linhas, mas as colunas têm que continuar as mesmas)."
        )

    decisoes = []
    for r in range(header_row + 1, ws.max_row + 1):
        tid = ws.cell(row=r, column=col_id).value
        data_planilha = ws.cell(row=r, column=col_data).value
        resposta = ws.cell(row=r, column=col_decisao).value
        if not tid or not resposta:
            continue
        resposta_norm = str(resposta).strip().lower()
        if resposta_norm in RESPOSTAS_SIM:
            decisoes.append((int(tid), 1, str(resposta).strip(), _normalizar_data(data_planilha)))
        elif resposta_norm in RESPOSTAS_NAO:
            decisoes.append((int(tid), 0, str(resposta).strip(), _normalizar_data(data_planilha)))
        else:
            print(f"  AVISO: linha {r} (chamado {tid}) tem resposta \"{resposta}\" não reconhecida (esperado Sim/Não) — ignorada.")
    return decisoes


def buscar_datas_reais(sreq):
    """ID -> Data real de abertura, direto do GLPI (busca leve, só ID e Data),
    pra conferir cada linha da planilha antes de mudar qualquer coisa."""
    params = [
        ("criteria[0][field]", "80"),
        ("criteria[0][searchtype]", "equals"),
        ("criteria[0][value]", str(ENTITY_OCORRENCIAS)),
        ("forcedisplay[0]", str(CAMPO_ID)),
        ("forcedisplay[1]", str(CAMPO_DATA)),
        ("range", "0-9999"),
    ]
    qs = urllib.parse.urlencode(params)
    res = sreq("GET", f"/search/Ticket?{qs}")
    datas = {}
    for row in res.get("data", []):
        tid = row.get(str(CAMPO_ID))
        if tid is None:
            continue
        datas[int(tid)] = _normalizar_data(row.get(str(CAMPO_DATA)))
    return datas


def main():
    if len(sys.argv) < 2:
        print("Uso: python importar_validacao_qualidade.py caminho_planilha_preenchida.xlsx")
        sys.exit(1)

    caminho = Path(sys.argv[1])
    if not caminho.exists():
        print(f"Arquivo não encontrado: {caminho}")
        sys.exit(1)

    cfg = json.loads((PASTA / "config.json").read_text(encoding="utf-8-sig"))
    base_url = cfg["api_url"]
    app_token = cfg["app_token"]
    user_token = cfg["user_token"]

    print(f"Lendo {caminho.name}...")
    decisoes = ler_planilha(caminho)
    print(f"{len(decisoes)} decisão(ões) reconhecida(s) na planilha.")
    if not decisoes:
        print("Nada pra importar.")
        return

    init = api_request(base_url, app_token, "GET", "/initSession", {"Authorization": "user_token " + user_token})
    session_token = init["session_token"]

    def sreq(method, path, data=None):
        return api_request(base_url, app_token, method, path, {"Session-Token": session_token}, data)

    try:
        print("Buscando registros de Classificação de todos os chamados (uma chamada só)...")
        todas_linhas_container6 = sreq("GET", "/PluginFieldsTicketdadosdaocorrncia?range=0-3000")
        por_ticket = {
            l["items_id"]: l for l in todas_linhas_container6
            if l.get("plugin_fields_containers_id") == CONTAINER_CLASSIFICACAO
        }

        print("Conferindo ID + Data de cada linha contra o GLPI...")
        datas_reais = buscar_datas_reais(sreq)

        estado = carregar_estado()
        agora = f"{datetime.now():%d/%m/%Y}"
        aplicados = 0
        ja_processados = 0
        pulados_sem_registro = 0
        pulados_data_divergente = 0

        for ticket_id, valor, resposta_original, data_planilha in decisoes:
            if str(ticket_id) in estado["importados"]:
                anterior = estado["importados"][str(ticket_id)]
                print(f"  chamado {ticket_id}: já processado antes ({anterior['data_importado']}, "
                      f"decisão \"{anterior['decisao']}\") — pulado, não reaplica.")
                ja_processados += 1
                continue

            data_real = datas_reais.get(ticket_id)
            if data_real is None:
                print(f"  chamado {ticket_id}: não encontrado na entidade Ocorrências no GLPI — pulado, confira o ID.")
                pulados_sem_registro += 1
                continue

            if data_planilha and data_real != data_planilha:
                print(f"  chamado {ticket_id}: AVISO — data da planilha ({data_planilha}) não bate com a "
                      f"data real no GLPI ({data_real}). Pulado por segurança, confira manualmente antes "
                      f"de reenviar essa linha.")
                pulados_data_divergente += 1
                continue

            registro = por_ticket.get(ticket_id)
            if not registro:
                print(f"  chamado {ticket_id}: sem registro de Classificação no GLPI — pulado.")
                pulados_sem_registro += 1
                continue

            sreq(
                "PUT",
                f"/PluginFieldsTicketdadosdaocorrncia/{registro['id']}",
                {"input": {
                    "id": registro["id"],
                    "avaliacaomotoristafield": valor,
                    "_disablenotifications": 1,
                    "_disablenotif": 1,
                }},
            )
            estado["importados"][str(ticket_id)] = {
                "data_importado": agora,
                "decisao": resposta_original,
                "arquivo": caminho.name,
            }
            aplicados += 1
            print(f"  chamado {ticket_id}: \"Inserir na avaliação motorista?\" = {resposta_original} (ID e Data conferidos)")

        salvar_estado(estado)
        print(f"\n{aplicados} chamado(s) atualizado(s) no GLPI.")
        if ja_processados:
            print(f"{ja_processados} já tinham sido processados antes (ignorados).")
        if pulados_data_divergente:
            print(f"{pulados_data_divergente} pulado(s) por divergência de data — confira essas linhas antes de reenviar.")
        if pulados_sem_registro:
            print(f"{pulados_sem_registro} pulado(s) por não ter registro correspondente no GLPI.")
        print(f"Estado salvo em {ARQ_ESTADO.name}.")
    finally:
        sreq("GET", "/killSession")


if __name__ == "__main__":
    main()
